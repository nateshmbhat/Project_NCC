import pandas as pd
from sqlite3 import connect, IntegrityError,OperationalError
from PyQt4.QtGui import QComboBox, QBrush, QPixmap, QApplication, QPrinter
from openpyxl import Workbook,load_workbook
from openpyxl.drawing.image import Image
import ENROLMENT_FORM
import os
from userinterface import Ui_MainWindow, _fromUtf8
from PyQt4 import QtCore, QtGui, QtWebKit
from shutil import copy2
from tempfile import TemporaryFile
from reportlab.lib.enums import TA_JUSTIFY,TA_CENTER,TA_LEFT,TA_RIGHT
from reportlab.lib.pagesizes import letter,A4
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer,Table,TableStyle,PageBreak
from reportlab.platypus import Image as pdfImage
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib import colors
from reportlab.lib.styles import ParagraphStyle
try:
    _fromUtf8 = QtCore.QString.fromUtf8
except AttributeError:
    def _fromUtf8(s):
        return s

try:
    _encoding = QtGui.QApplication.UnicodeUTF8


    def _translate(context, text, disambig):
        return QtGui.QApplication.translate(context, text, disambig, _encoding)
except AttributeError:
    def _translate(context, text, disambig):
        return QtGui.QApplication.translate(context, text, disambig)


class Ui_loginDialog(object):
    def setupUi(self, loginDialog):
        loginDialog.setObjectName(_fromUtf8("loginDialog"))
        loginDialog.setWindowModality(QtCore.Qt.ApplicationModal)
        loginDialog.resize(616, 309)
        sizePolicy = QtGui.QSizePolicy(QtGui.QSizePolicy.Fixed, QtGui.QSizePolicy.Fixed)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(loginDialog.sizePolicy().hasHeightForWidth())
        loginDialog.setSizePolicy(sizePolicy)
        loginDialog.setMinimumSize(QtCore.QSize(616, 298))
        loginDialog.setMaximumSize(QtCore.QSize(876, 411))
        font = QtGui.QFont()
        font.setFamily(_fromUtf8("Centaur"))
        font.setPointSize(10)
        font.setBold(True)
        font.setWeight(75)
        loginDialog.setFont(font)
        icon = QtGui.QIcon()
        icon.addPixmap(QtGui.QPixmap(_fromUtf8(":/icons/ncc2.png")), QtGui.QIcon.Normal, QtGui.QIcon.Off)
        loginDialog.setWindowIcon(icon)
        loginDialog.setStyleSheet(_fromUtf8("#loginDialog{\n"
                                            "border-image:url(:/icons/image-blur.png);\n"
                                            "background-position:center;\n"
                                            "}"))
        loginDialog.setSizeGripEnabled(True)
        self.verticalLayout = QtGui.QVBoxLayout(loginDialog)
        self.verticalLayout.setObjectName(_fromUtf8("verticalLayout"))
        self.horizontalLayout = QtGui.QHBoxLayout()
        self.horizontalLayout.setObjectName(_fromUtf8("horizontalLayout"))
        self.label = QtGui.QLabel(loginDialog)
        sizePolicy = QtGui.QSizePolicy(QtGui.QSizePolicy.Fixed, QtGui.QSizePolicy.Preferred)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.label.sizePolicy().hasHeightForWidth())
        self.label.setSizePolicy(sizePolicy)
        self.label.setMinimumSize(QtCore.QSize(250, 96))
        font = QtGui.QFont()
        font.setFamily(_fromUtf8("Centaur"))
        font.setPointSize(20)
        font.setBold(True)
        font.setItalic(False)
        font.setWeight(75)
        self.label.setFont(font)
        self.label.setStyleSheet(_fromUtf8("\n"
                                           "color:white;"))
        self.label.setAlignment(QtCore.Qt.AlignCenter)
        self.label.setObjectName(_fromUtf8("label"))
        self.horizontalLayout.addWidget(self.label)
        self.login_usernameLineEdit = QtGui.QLineEdit(loginDialog)
        sizePolicy = QtGui.QSizePolicy(QtGui.QSizePolicy.Preferred, QtGui.QSizePolicy.Fixed)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.login_usernameLineEdit.sizePolicy().hasHeightForWidth())
        self.login_usernameLineEdit.setSizePolicy(sizePolicy)
        self.login_usernameLineEdit.setMinimumSize(QtCore.QSize(0, 30))
        font = QtGui.QFont()
        font.setFamily(_fromUtf8("Century"))
        font.setPointSize(15)
        font.setBold(True)
        font.setWeight(75)
        self.login_usernameLineEdit.setFont(font)
        self.login_usernameLineEdit.setStyleSheet(_fromUtf8("border-radius:6px;\n"
                                                            "border:1px groove purple;\n"
                                                            "\n"
                                                            "margin-right:10px;"))
        self.login_usernameLineEdit.setObjectName(_fromUtf8("login_usernameLineEdit"))
        self.horizontalLayout.addWidget(self.login_usernameLineEdit)
        self.verticalLayout.addLayout(self.horizontalLayout)
        self.horizontalLayout_2 = QtGui.QHBoxLayout()
        self.horizontalLayout_2.setObjectName(_fromUtf8("horizontalLayout_2"))
        self.label_2 = QtGui.QLabel(loginDialog)
        self.label_2.setMinimumSize(QtCore.QSize(250, 96))
        self.label_2.setStyleSheet(_fromUtf8("font: 75 20pt \"Centaur\";\n"
                                             "font-weight:bold;\n"
                                             "color:white;"))
        self.label_2.setAlignment(QtCore.Qt.AlignCenter)
        self.label_2.setObjectName(_fromUtf8("label_2"))
        self.horizontalLayout_2.addWidget(self.label_2)
        self.login_passwordLineEdit = QtGui.QLineEdit(loginDialog)
        self.login_passwordLineEdit.setMinimumSize(QtCore.QSize(0, 30))
        font = QtGui.QFont()
        font.setFamily(_fromUtf8("Century"))
        font.setPointSize(15)
        font.setBold(True)
        font.setWeight(75)
        self.login_passwordLineEdit.setFont(font)
        self.login_passwordLineEdit.setStyleSheet(_fromUtf8("border-radius:6px;\n"
                                                            "border:1px groove purple;\n"
                                                            "margin-right:10px;"))
        self.login_passwordLineEdit.setEchoMode(QtGui.QLineEdit.Password)
        self.login_passwordLineEdit.setObjectName(_fromUtf8("login_passwordLineEdit"))
        self.horizontalLayout_2.addWidget(self.login_passwordLineEdit)
        self.verticalLayout.addLayout(self.horizontalLayout_2)
        self.loginPushButton = QtGui.QPushButton(loginDialog)
        font = QtGui.QFont()
        font.setFamily(_fromUtf8("Longdon Decorative"))
        font.setPointSize(29)
        font.setBold(False)
        font.setWeight(50)
        self.loginPushButton.setFont(font)
        self.loginPushButton.setStyleSheet(_fromUtf8("#loginPushButton\n"
                                                     "{\n"
                                                     "background-color:qlineargradient(spread:pad, x1:0, y1:0, x2:1, y2:0, stop:0 rgba(26, 41, 24, 230), stop:0.085 rgba(2, 79, 0, 255), stop:0.221591 rgba(50, 147, 22, 255), stop:0.275 rgba(165, 142, 70, 255), stop:0.431818 rgba(243, 100, 79, 255), stop:0.573864 rgba(135, 95, 80, 255), stop:0.667 rgba(137, 97, 255, 255), stop:0.818182 rgba(160, 255, 244, 255), stop:0.885 rgba(193, 222, 185, 255), stop:1 rgba(93, 128, 0, 255));\n"
                                                     "color:white;\n"
                                                     "border-radius:10px;\n"
                                                     "border-top:1px solid yellow;\n"
                                                     "border-bottom:1px solid black;\n"
                                                     "padding:0px 0px;\n"
                                                     "}\n"
                                                     "\n"
                                                     "#loginPushButton:hover{\n"
                                                     "color:black;\n"
                                                     "    background-color:rgba(0, 170, 255,255);\n"
                                                     "\n"
                                                     "}\n"
                                                     "\n"
                                                     "#loginPushButton:pressed{\n"
                                                     "color:black;\n"
                                                     "    background-color:rgba(0, 170, 255,175);\n"
                                                     "\n"
                                                     "}"))
        self.loginPushButton.setFlat(True)
        self.loginPushButton.setObjectName(_fromUtf8("loginPushButton"))
        self.verticalLayout.addWidget(self.loginPushButton, QtCore.Qt.AlignBottom)
        self.label_3 = QtGui.QLabel(loginDialog)
        sizePolicy = QtGui.QSizePolicy(QtGui.QSizePolicy.Minimum, QtGui.QSizePolicy.Preferred)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.label_3.sizePolicy().hasHeightForWidth())
        self.label_3.setSizePolicy(sizePolicy)
        self.label_3.setMinimumSize(QtCore.QSize(0, 0))
        self.label_3.setMaximumSize(QtCore.QSize(16777215, 21))
        font = QtGui.QFont()
        font.setFamily(_fromUtf8("MS Shell Dlg 2"))
        font.setPointSize(7)
        self.label_3.setFont(font)
        self.label_3.setObjectName(_fromUtf8("label_3"))
        self.verticalLayout.addWidget(self.label_3, QtCore.Qt.AlignBottom)

        self.retranslateUi(loginDialog)
        QtCore.QMetaObject.connectSlotsByName(loginDialog)

    def retranslateUi(self, loginDialog):
        loginDialog.setWindowTitle(_translate("loginDialog", "LOGIN", None))
        self.label.setText(_translate("loginDialog",
                                      "<html><head/><body><p><span style=\" font-size:22pt; color:#fafde8;\">Username</span></p></body></html>",
                                      None))
        self.label_2.setText(_translate("loginDialog", "Password", None))
        self.loginPushButton.setText(_translate("loginDialog", "Login", None))
        self.label_3.setText(_translate("loginDialog",
                                        "Developed by 4Karnataka Battalion, NCC  [ by Natesh M Bhat and Gangadhara Shetty P J ]",
                                        None))


class logic():
    flag = 0

    def __init__(self):

        ENROLMENT_FORM.enroll().create_table_Attendance()

        ENROLMENT_FORM.enroll().create_table_marks_A_cert()

        ENROLMENT_FORM.enroll().create_table_marks_B_cert()

        ENROLMENT_FORM.enroll().create_table_marks_C_cert()

        ENROLMENT_FORM.enroll().create_table_Enrolment()

        ENROLMENT_FORM.enroll().create_table_camps()



        ui.NullcertRadioButton.setChecked(True)

        ui.vegRadioButton.setChecked(True)

        ui.insertcondition.clicked.connect(self.coninsert)

        ui.clearcondition.clicked.connect(lambda: ui.conditionsentrylabel.setText(""))

        ui.backcondition.clicked.connect(self.conback)

        ui.andcondition.clicked.connect(lambda: ui.conditionsentrylabel.setText(ui.conditionsentrylabel.text().strip() + " and "))

        ui.orcondition.clicked.connect(
            lambda: ui.conditionsentrylabel.setText(ui.conditionsentrylabel.text().strip() + " or "))

        ui.openbracecondition.clicked.connect(
            lambda: ui.conditionsentrylabel.setText(ui.conditionsentrylabel.text().strip() + "("))

        ui.closebracecondition.clicked.connect(
            lambda: ui.conditionsentrylabel.setText(ui.conditionsentrylabel.text().strip() + ")"))

        ui.querycondition.clicked.connect(self.conquery)

        ui.submitPushButton.clicked.connect(self.check_enrol_form_data)

        ui.enrolPushButton.pressed.connect(self.enrol_button_pressed)

        ui.aadhaarnumRadioButton.toggled.connect(self.enrol_adhaar_radio_change)

        self.sql = ''

        self.candidphoto = ''

        self.signaturephoto = ''

        ui.selectpicturePushButton.clicked.connect(lambda: self.picselect(ui.selectpicturePushButton))

        ui.startdateDateEdit.hide()

        ui.startdateLabel.hide()

        ui.enddateDateEdit.hide()

        ui.enddateLabel.hide()

        ui.enrolmentuploaddataLineEdit.hide()

        ui.locationLineEdit.hide()

        ui.campsNameuploaddataComboBox.hide()

        ui.bloodgroupqueryComboBox.hide()

        ui.institutionqueryComboBox.hide()

        ui.seniorityqueryComboBox.hide()

        ui.rankqueryComboBox.hide()

        ui.sexqueryComboBox.hide()

        ui.datequeryDateEdit.hide()

        ui.certificatequeryComboBox.hide()

        ui.vegitarianqueryComboBox.hide()

        ui.campsattendedqueryComboBox.hide()

        ui.mobileLineEdit.setValidator(QtGui.QDoubleValidator())

        ui.accountnumLineEdit.setValidator(QtGui.QDoubleValidator())

        ui.micrLineEdit.setValidator(QtGui.QDoubleValidator())

        ui.micrLineEdit.setValidator(QtGui.QDoubleValidator())

        ui.aadhaarLineEdit.setValidator(QtGui.QDoubleValidator())

        ui.marksLineEdit.setValidator(QtGui.QDoubleValidator())

        ui.vegRadioButton.setChecked(True)

        ui.NullcertRadioButton.setChecked(True)

        ui.enrolmentCheckBox.stateChanged.connect(lambda: self.querycheckboxes(ui.enrolmentCheckBox))

        ui.searchPushButton.clicked.connect(lambda: self.display(ui.searchPushButton))

        ui.rankCheckBox.stateChanged.connect(lambda: self.querycheckboxes(ui.rankCheckBox))

        ui.aadhaarCheckBox.stateChanged.connect(lambda: self.querycheckboxes(ui.aadhaarCheckBox))

        ui.sfnameCheckBox.stateChanged.connect(lambda: self.querycheckboxes(ui.sfnameCheckBox))

        ui.ffnameCheckBox.stateChanged.connect(lambda: self.querycheckboxes(ui.ffnameCheckBox))

        ui.mfnameCheckBox.stateChanged.connect(lambda: self.querycheckboxes(ui.mfnameCheckBox))

        ui.sexCheckBox.stateChanged.connect(lambda: self.querycheckboxes(ui.sexCheckBox))

        ui.bloodgroupCheckBox.stateChanged.connect(lambda: self.querycheckboxes(ui.bloodgroupCheckBox))

        ui.emailCheckBox.stateChanged.connect(lambda: self.querycheckboxes(ui.emailCheckBox))

        ui.mobileCheckBox.stateChanged.connect(lambda: self.querycheckboxes(ui.mobileCheckBox))

        ui.addressCheckBox.stateChanged.connect(lambda: self.querycheckboxes(ui.addressCheckBox))

        ui.dateofbirthCheckBox.stateChanged.connect(lambda: self.querycheckboxes(ui.dateofbirthCheckBox))

        ui.institutionCheckBox.stateChanged.connect(lambda: self.querycheckboxes(ui.institutionCheckBox))

        ui.unitCheckBox.stateChanged.connect(lambda: self.querycheckboxes(ui.unitCheckBox))

        ui.banknameCheckBox.stateChanged.connect(lambda: self.querycheckboxes(ui.banknameCheckBox))

        ui.bankbranchCheckBox.stateChanged.connect(lambda: self.querycheckboxes(ui.bankbranchCheckBox))

        ui.accountnameCheckBox.stateChanged.connect(lambda: self.querycheckboxes(ui.accountnameCheckBox))

        ui.accountnumCheckBox.stateChanged.connect(lambda: self.querycheckboxes(ui.accountnumCheckBox))

        ui.ifsccodeCheckBox.stateChanged.connect(lambda: self.querycheckboxes(ui.ifsccodeCheckBox))

        ui.policestationCheckBox.stateChanged.connect(lambda: self.querycheckboxes(ui.policestationCheckBox))

        ui.educationCheckBox.stateChanged.connect(lambda: self.querycheckboxes(ui.educationCheckBox))

        ui.earliercandidateCheckBox.stateChanged.connect(lambda: self.querycheckboxes(ui.earliercandidateCheckBox))


        ui.pannumCheckBox.stateChanged.connect(lambda: self.querycheckboxes(ui.pannumCheckBox))

        ui.schoolcollegeCheckBox.stateChanged.connect(lambda: self.querycheckboxes(ui.schoolcollegeCheckBox))

        ui.seniorityCheckBox.stateChanged.connect(lambda: self.querycheckboxes(ui.seniorityCheckBox))

        ui.specialAchievementsCheckBox.stateChanged.connect(
            lambda: self.querycheckboxes(ui.specialAchievementsCheckBox))

        ui.extraCurricularActivitiesCheckBox.stateChanged.connect(
            lambda: self.querycheckboxes(ui.extraCurricularActivitiesCheckBox))

        ui.remarksCheckBox.stateChanged.connect(lambda: self.querycheckboxes(ui.remarksCheckBox))

        ui.micrCheckBox.stateChanged.connect(lambda: self.querycheckboxes(ui.micrCheckBox))

        ui.enrollDateCheckBox.stateChanged.connect(lambda: self.querycheckboxes(ui.enrollDateCheckBox))

        ui.campsAttendedCheckBox.stateChanged.connect(lambda: self.querycheckboxes(ui.campsAttendedCheckBox))

        ui.vegitarianCheckBox.stateChanged.connect(lambda: self.querycheckboxes(ui.vegitarianCheckBox))

        ui.selectallCheckBox.stateChanged.connect(self.queryselectall)



        ui.saveExelPushButton.clicked.connect(self.saveExcelfuntion)

        ui.vegRadioButton.setChecked(True)

        ui.NullcertRadioButton.setChecked(True)

        if not os.path.exists(r'candidate photos'):
            os.mkdir(r'candidate photos')

        ui.savedataPushButton.clicked.connect(self.saveuploadeddata)

        ui.save_data_excelPushButton.clicked.connect(self.saveexceluploadeddata)

        ui.updateExelPushButton.clicked.connect(self.update_excel_function)

        ui.settings_addinstitutionPushButton.clicked.connect(lambda: (
            ui.settings_addinstitutionPushButton.hide(), ui.removeinstitutionPushButton.hide(),
            ui.settings_instLineEdit.show(), ui.settings_addPushButton.show(), ui.settings_backinstPushButton.show()))

        ui.settings_addPushButton.clicked.connect(lambda: self.institution_add_or_remove(ui.settings_addPushButton))

        ui.removeinstitutionPushButton.clicked.connect(
            lambda: self.institution_add_or_remove(ui.removeinstitutionPushButton))

        ui.settings_backinstPushButton.clicked.connect(lambda: self.set_institutions_list())

        ui.generateexcelqueryPushButton.clicked.connect(self.generateexcelforquery)



        ENROLMENT_FORM.enroll().create_table_camps()

        ui.startdateDateEdit.hide()

        ui.startdateLabel.hide()

        ui.enddateDateEdit.hide()

        ui.enddateLabel.hide()

        ui.enrolmentuploaddataLineEdit.hide()

        ui.locationLineEdit.hide()

        ui.campsNameuploaddataComboBox.hide()

        ui.showlongnrPushButton.clicked.connect(self.showlongnr)


        ui.institutionlongnrComboBox.hide()
        ui.unitlongnrLineEdit.hide()
        ui.generateexcellongnrPushButton.clicked.connect(self.generateExcelForLongNr)
        ui.updateexcellongnrPushButton.clicked.connect(self.updateExcelForLongNr)
        ui.certificateComboBox.hide()

        ui.yearComboBox.hide()
        ui.earlierenrolmentnumLineEdit.hide()
        ui.eligibilityCheckBox.hide()
        ui.eligibilityCheckBox.stateChanged.connect(self.eligibilitylogic)
        ui.query_backPushButton.clicked.connect(lambda: (self.show_query_elements(), ui.query_backPushButton.hide()))

        ui.helpWebView.load(QtCore.QUrl('Help/help.html'))

        self.view = QtWebKit.QWebView()
        self.printer = QPrinter()
        self.printer.setOutputFormat(QPrinter.PdfFormat)
        self.printer.setPageMargins(0.5, 0.5, 0.5, 0.5, QPrinter.Inch)
        self.printer.setPaperSize(QPrinter.A4)
        self.init_settings()
        ui.earliercandidateComboBox.currentIndexChanged.connect(self.markslogic)
        ui.conditionlistcombobox.currentIndexChanged.connect(self.conditionscomboboxlogic)
        ui.institutionuploaddatacomboBox.currentIndexChanged.connect(self.openuploaddata)
        ui.certificateComboBox.currentIndexChanged.connect(self.openuploaddata)
        ui.yearComboBox.currentIndexChanged.connect(self.openuploaddata)
        ui.campsNameuploaddataComboBox.currentIndexChanged.connect(self.openuploaddata)
        ui.typelongnrComboBox.currentIndexChanged.connect(self.typelongNrComboBoxLogic)
        ui.typecomboBox.currentIndexChanged.connect(self.typecomboboxlogic)

    def markslogic(self):
        if ui.earliercandidateComboBox.currentText()=="Yes":
            ui.earlierenrolmentnumLineEdit.show()
        else:
            ui.earlierenrolmentnumLineEdit.hide()
    def init_settings(self):

        loginui.firstrun = False

        """Sets all the parameters from the settings file"""

        self.settings = QtCore.QSettings('settings.ini', QtCore.QSettings.IniFormat)
        self.institutionlist = self.settings.value('institutionlist').split(',,,')
        self.set_institutions_list()

        '''ALL STANDARD FIELDS LIST (Fields that are in the master enrolment form )'''

        self.all_enrolment_form_fieldslist = self.settings.value('all_enrolment_form_fieldslist').split(',,,')
        ui.mytab.setCurrentIndex(int(self.settings.value('current_tab')))
        geo = [int(i) for i in self.settings.value('window_geometry').split(',,,')]

        MainWindow.setGeometry(geo[0], geo[1], geo[2], geo[3])
        app.aboutToQuit.connect(self.handler)

        """List of all forms in the forms tab"""
        self.formslist = self.settings.value('formslist').strip().split(',,,');
        ui.formsComboBox.clear()
        ui.formsComboBox.addItems(self.formslist)

        '''This is the list of all the types of entry forms in data entry tab'''
        self.dataentrytypes = self.settings.value('dataentrytypes').strip().split(',,,')
        ui.typecomboBox.clear()
        ui.typecomboBox.addItems(self.dataentrytypes)

        '''List of forms and fields in the settings tab'''
        self.formslist = self.settings.value('formslist').strip().split(',,,')
        self.clear = ui.settings_formsListWidget.clear()

        ui.settings_formsListWidget.setSpacing(1)
        self.set_forms_list()

        '''List of all the fields under Dataentry Types '''

        self.marksA_fieldslist = self.settings.value('marksA_fieldslist').split(',,,')
        self.marksB_fieldslist = self.settings.value('marksB_fieldslist').split(',,,')
        self.marksC_fieldslist = self.settings.value('marksC_fieldslist').split(',,,')
        self.camps_fieldslist = self.settings.value('camps_fieldslist').split(',,,')
        self.extraactivities_fieldslist = self.settings.value('extraactivities_fieldslist').split(',,,')
        self.remarks_fieldslist = self.settings.value('remarks_fieldslist').split(',,,')

        # hiding fields in the settigns_fieldssectoin
        ui.settings_addfieldPushButton.hide()
        ui.settings_addfieldLineEdit.hide()
        ui.settings_fieldsComboBox.hide()
        ui.settings_fieldsknownRadioButton.hide()
        ui.settings_fieldsunknownRadioButton.hide()
        ui.settings_removefieldPushButton.hide()
        ui.settings_removeformPushButton.hide()
        ui.query_backPushButton.hide()

        self.nametolistsql = {}
        self.nametolistnotsql = {}

        for form in self.formslist:
            sqlfieldlist = self.settings.value(form.replace(' ', '_') + '_sql_fieldslist')
            notsqlfieldlist = self.settings.value(form.replace(' ', '_') + '_notsql_fieldslist')

            sqlfieldlist = sqlfieldlist.split(',,,') if sqlfieldlist else []
            notsqlfieldlist = notsqlfieldlist.split(',,,') if notsqlfieldlist else []

            self.nametolistsql.update({form: sqlfieldlist})
            self.nametolistnotsql.update({form: notsqlfieldlist})

        # The below lines are used to connect the widgets to the corresponding functions

        ui.settings_addformPushButton.clicked.connect(
            lambda: self.settings_form_field_add_remove(ui.settings_addformPushButton))
        ui.settings_addfieldPushButton.clicked.connect(
            lambda: self.settings_form_field_add_remove(ui.settings_addfieldPushButton))
        ui.settings_formsListWidget.itemClicked.connect(self.settings_form_item_clicked)

        ui.settings_removefieldPushButton.clicked.connect(
            lambda: self.settings_form_field_add_remove(ui.settings_removefieldPushButton))

        ui.settings_removeformPushButton.clicked.connect(
            lambda: self.settings_form_field_add_remove(ui.settings_removeformPushButton))

        for i in ui.Enrol.findChildren((QtGui.QLineEdit, QtGui.QComboBox, QtGui.QDateEdit, QtGui.QTextEdit)):
            if i.objectName() == 'searchbyfieldLineEdit':
                continue
            s = '#' + i.objectName() + ''':focus
                    {
                        border:2px groove chartreuse;
                    }'''
            i.setStyleSheet(s)

        ui.settings_candidopenPushButton.clicked.connect(lambda: os.system('start explorer "candidate photos"'))
        ui.enroldateDateEdit.setDate(QtCore.QDate.currentDate())

        ui.settings_removecampPushButton.hide()
        ui.exceltodatabasePushButton.clicked.connect(self.save_from_excel_to_database)

        def camplist_clicked():
            if ui.settings_campslistListWidget.currentItem().text().strip() in ['NIC', 'CATC', 'AAC', 'Mounaineering',
                                                                                'Trekking', 'SSB', 'BLC', 'ALC', 'RDC',
                                                                                'TSC', 'Snow Skiing']:
                ui.settings_removecampPushButton.hide()
            else:
                ui.settings_removecampPushButton.show()

        '''CAMPS LIST'''
        self.allcampslist = self.settings.value('allcampslist').split(',,,')
        self.set_camps_list()

        ui.settings_addcampPushButton.clicked.connect(lambda: self.add_remove_camp(ui.settings_addcampPushButton))

        ui.settings_removecampPushButton.clicked.connect(lambda: self.add_remove_camp(ui.settings_removecampPushButton))
        ui.settings_campslistListWidget.itemClicked.connect(camplist_clicked)
        ui.settings_backupdataPushButton.clicked.connect(self.backupdata)
        ui.settings_restoredataPushButton.clicked.connect(self.restoredata)

        # -----------------------------------------------------------------------------------------------------------------------------------------

        ui.enrol_signaturePushButton.clicked.connect(lambda: self.picselect(ui.enrol_signaturePushButton))

        def settings_login_press():
            loginDialog.show()

        ui.settings_loginPushButton.clicked.connect(settings_login_press)

        self.login_permission()
        self.showtooltip("WELCOME")




    def query_hide_elements(self):
        ui.checkboxFrame.hide()
        ui.conditionsentrylabel.hide()

        for i in ui.frame.findChildren((QtGui.QPushButton, QtGui.QLineEdit, QtGui.QComboBox, QtGui.QDateEdit)):
            i.hide()
        ui.query_backPushButton.show()

    def show_query_elements(self):
        ui.checkboxFrame.show()
        ui.conditionsentrylabel.show()

        for i in ui.frame.findChildren((QtGui.QPushButton, QtGui.QLineEdit)):
            i.show()
        ui.valuelineEdit.show()
        ui.conditionlistcombobox.show()
        ui.conditionlistcombobox.setCurrentIndex(0)

    def login_permission(self):
        app.processEvents()

        if loginui.username == 'ncc_viewer':
            ui.savedataPushButton.setDisabled(True)
            ui.updateentryCheckBox.hide()
            ui.submitPushButton.hide()
            ui.enrolPushButton.setDisabled(True)

            for i in ui.Settings.findChildren((QtGui.QPushButton, QtGui.QLineEdit)):
                if i.text() != "Open Candidates Picture folder" and i.text() != 'LOGIN':
                    i.setDisabled(True)

        if loginui.username == 'ncc_editor':
            ui.savedataPushButton.setDisabled(False)
            ui.updateentryCheckBox.show()
            ui.submitPushButton.show()
            ui.enrolPushButton.setDisabled(False)

            for i in ui.Settings.findChildren((QtGui.QPushButton, QtGui.QLineEdit)):
                if i.text() != "Open Candidates Picture folder":
                    i.setDisabled(False)

    def check_if_img_exists(self, imagename):
        if os.path.exists(r'candidate photos\{}.png'.format(imagename)):
            img = r'candidate photos\{}.png'.format(imagename)

        elif os.path.exists(r'candidate photos\{}.jpg'.format(imagename)):
            img = r'candidate photos\{}.jpg'.format(imagename)

        elif os.path.exists(r'candidate photos\{}.JPG'.format(imagename)):
            img = r'candidate photos\{}.JPG'.format(imagename)

        elif os.path.exists(r'candidate photos\{}.PNG'.format(imagename)):
            img = r'candidate photos\{}.PNG'.format(imagename)
        else:
            img = ''

        return img

    columnnameinexcel = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S',
                         'T', 'U', 'V', 'W', 'X', 'Y', 'Z', 'AA', 'AB', 'AC', 'AD', 'AE', 'AF', 'AG', 'AH', 'AI', 'AJ',
                         'AK', 'AL', 'AM', 'AN','AO','AP','AQ','AR','AS','AT','AU','AV','AW','AX','AY','AZ','BA','BB','BC','BD']



    def generateExcelForLongNr(self):
        book = Workbook()
        sheet = book.active
        headingdata = self.settings.value("enrolmentfields").split(',,,')
        headingdata.insert(0, "Photo")
        headingdata.insert(1, "Signature")
        sheet.append(headingdata)
        name = QtGui.QFileDialog.getSaveFileName(directory=r"C:\Users\{}\Documents".format(os.getlogin()),
                                                 caption="Save File",
                                                 filter=".xlsx")
        if not len(name):
            return
        columnwidth = []
        for i in range(ui.tableWidget_2.rowCount()):
            for j in range(len(headingdata)):
                if j == 0:
                    candidatephoto = self.check_if_img_exists(ui.tableWidget_2.item(i, j + 2).text())

                    if candidatephoto:
                        img = Image(candidatephoto, size=[125, 180])
                        img.anchor(sheet['A' + str(i + 2)])
                        sheet.add_image(img)

                    sheet.row_dimensions[i + 2].height = 80
                    sheet.column_dimensions["A"].width = 18
                elif j == 1:
                    candidatephotosign = self.check_if_img_exists(ui.tableWidget_2.item(i, j + 1).text() + "_sign")
                    if candidatephotosign:
                        img1 = Image(candidatephotosign, size=[125, 180])
                        img1.anchor(sheet['B' + str(i + 2)])
                        img1.drawing.left = 127
                        sheet.add_image(img1)
                    sheet.row_dimensions[i + 2].height = 80
                    sheet.column_dimensions["B"].width = 18
                else:
                    sheet.cell(row=i + 2, column=j + 1).value = ui.tableWidget_2.item(i, j).text()
        for i in range(ui.tableWidget_2.columnCount()):
            columnwidth.append(int(ui.tableWidget_2.columnWidth(i) / 10 + 5))
        for i in range(len(columnwidth)):
            sheet.column_dimensions[self.columnnameinexcel[i + 2]].width = columnwidth[i]
        book.save(name)
        book.save(TemporaryFile())
        self.showtooltip("Excel file created sucessfully")
        os.startfile(name)





    def updateExcelForLongNr(self):

        name = QtGui.QFileDialog.getOpenFileName(directory=r"C:\Users\{}\Documents".format(os.getlogin()),
                                                 caption="Save File")
        if not len(name):
            return
        book =load_workbook(name)
        sheet = book.active
        columnwidth = []
        oldrows = sheet.max_row
        for i in range(oldrows):
            candidatephoto = self.check_if_img_exists(sheet.cell(row=i + 1, column=3).value)
            if candidatephoto:
                img = Image(candidatephoto, size=[125, 180])
                img.anchor(sheet['A' + str(i + 1)])
                sheet.add_image(img)

            sheet.row_dimensions[i + 2].height = 80
            sheet.column_dimensions["A"].width = 18
            candidatephotosign = self.check_if_img_exists(sheet.cell(row=i + 1, column=3).value + "_sign")
            if candidatephotosign:
                img1 = Image(candidatephotosign, size=[125, 180])
                img1.anchor(sheet['B' + str(i + 1)])
                img1.drawing.left = 127
                sheet.add_image(img1)
            sheet.row_dimensions[i + 2].height = 80
            sheet.column_dimensions["B"].width = 18
        for i in range(ui.tableWidget_2.rowCount()):
            for j in range(ui.tableWidget_2.columnCount()):
                if j == 0:
                    candidatephoto = self.check_if_img_exists(ui.tableWidget_2.item(i, j + 2).text())
                    if candidatephoto:
                        img = Image(candidatephoto, size=[125, 180])
                        img.anchor(sheet['A' + str(i + 1 + oldrows)])
                        sheet.add_image(img)

                    sheet.row_dimensions[i + 1 + oldrows].height = 80
                    sheet.column_dimensions["A"].width = 18
                elif j == 1:
                    candidatephotosign = self.check_if_img_exists(ui.tableWidget_2.item(i, j + 1).text() + "_sign")
                    if candidatephotosign:
                        img1 = Image(candidatephotosign, size=[125, 180])
                        img1.anchor(sheet['B' + str(i + 1 + oldrows)])
                        img1.drawing.left = 127
                        sheet.add_image(img1)
                    sheet.row_dimensions[i + 1 + oldrows].height = 80
                    sheet.column_dimensions["B"].width = 18
                else:
                    sheet.cell(row=i + 1 + oldrows, column=j + 1).value = ui.tableWidget_2.item(i, j).text()

        for i in range(ui.tableWidget_2.columnCount()):
            columnwidth.append(int(ui.tableWidget_2.columnWidth(i) / 10 + 5))
        for i in range(len(columnwidth)):
            sheet.column_dimensions[self.columnnameinexcel[i + 2]].width = columnwidth[i]
        book.save(name)
        book.save(TemporaryFile())
        self.showtooltip("Excel file created sucessfully")
        os.startfile(name)




    def showlongnr(self):
        ui.tableWidget_2.setSelectionMode(QtGui.QAbstractItemView.NoSelection)
        ui.tableWidget_2.setRowCount(0)
        ui.tableWidget_2.setColumnCount(0)
        selectiontype = ui.typelongnrComboBox.currentText()
        selectedinstitution = ui.institutionlongnrComboBox.currentText()
        selectedunit = ui.unitlongnrLineEdit.text()
        sqldata = []
        """ui.tableWidget_2.setCellWidget(i,j, ImgWidget1(self))"""
        if selectiontype == "Institution":
            sqldata = ENROLMENT_FORM.enroll().execute(
                "select * from enrolment where Institution='" + selectedinstitution + "'")
        elif selectiontype == "Unit":
            sqldata = ENROLMENT_FORM.enroll().execute(
                "select * from enrolment where Unit LIKE '" + selectedunit + "' collate utf8_general_ci")
        elif selectiontype == "All":
            sqldata = ENROLMENT_FORM.enroll().execute("select * from enrolment")
        if not len(sqldata) > 0:
            ui.tableWidget.clear()
            self.showtooltip("No data found")

            ui.tableWidget_2.setColumnCount(0)
            ui.tableWidget_2.setRowCount(0)
            return

        verticalheader = []
        horizontalheader = list(self.settings.value("enrolmentfields").split(',,,'))
        horizontalheader.insert(0, "Photo")
        horizontalheader.insert(1, "Signature")
        for i in range(len(sqldata)):
            verticalheader.append(sqldata[i][0])
        ui.tableWidget_2.setColumnCount(len(horizontalheader))
        ui.tableWidget_2.setRowCount(len(verticalheader))
        ui.tableWidget_2.setHorizontalHeaderLabels(horizontalheader)
        ui.tableWidget_2.setVerticalHeaderLabels(verticalheader)
        for i in range(len(sqldata)):
            for j in range(len(horizontalheader)):
                if j == 0:
                    self.imagePath = self.check_if_img_exists(sqldata[i][j])
                    if self.imagePath:
                        pixmap = QtGui.QPixmap(self.imagePath).scaled(115, 120)
                        qbimg = QBrush(pixmap)
                        item = QtGui.QTableWidgetItem()
                        item.setBackground(qbimg)
                        ui.tableWidget_2.setItem(i, j, item)
                elif j == 1:
                    self.imagePath = self.check_if_img_exists(sqldata[i][j - 1] + "_sign")
                    if self.imagePath:
                        # pixmap = QtGui.QPixmap(self.imagePath).scaled(135, 120)
                        # qbimg = QBrush(pixmap)
                        item = QtGui.QTableWidgetItem()
                        item.setBackground(QBrush(QPixmap(self.imagePath).scaled(115, 120)))
                        ui.tableWidget_2.setItem(i, j, item)
                else:
                    ui.tableWidget_2.setItem(i, j, QtGui.QTableWidgetItem(sqldata[i][j - 2]))
        myfont = QtGui.QFont()
        myfont.setBold(True)
        myfont.setFamily("georgia")
        ui.tableWidget_2.resizeColumnsToContents()
        ui.tableWidget_2.resizeColumnsToContents()
        for i in range(ui.tableWidget_2.rowCount()):
            ui.tableWidget_2.setRowHeight(i, 120)

            for j in range(ui.tableWidget_2.columnCount()):
                if j != 0 and j != 1:
                    ui.tableWidget_2.item(i, j).setBackground(QtGui.QColor(150, 150, 150,200))
                    ui.tableWidget_2.item(i, j).setFont(myfont)
                    ui.tableWidget_2.item(i, j).setTextAlignment(QtCore.Qt.AlignCenter)
                ui.tableWidget_2.horizontalHeader().setResizeMode(j, QtGui.QHeaderView.ResizeToContents)
            ui.tableWidget_2.verticalHeader().setResizeMode(i, QtGui.QHeaderView.Fixed)

        ui.tableWidget_2.showGrid()

        ui.tableWidget_2.setStyleSheet(
            "color:white;font-weight:bold;font-size:18px;background-color:transparent;border:1px solid black;gridline-color:black;")
        ui.tableWidget_2.horizontalHeader().setStyleSheet(
            "color:darkgreen;font-size:25px;font-weight:bold;font-family:gabriola;border:1px solid black;")
        ui.tableWidget_2.verticalHeader().setStyleSheet(
            "color:darkorange;width:50px;font-size:30px;font-family:gabriola;border:1px solid black;gridline-color:black;")
        ui.tableWidget_2.setEditTriggers(QtGui.QAbstractItemView.NoEditTriggers)


    def typelongNrComboBoxLogic(self):
        selectiontype = ui.typelongnrComboBox.currentText()
        if selectiontype == "Selection By" or selectiontype == "All":
            ui.institutionlongnrComboBox.hide()
            ui.unitlongnrLineEdit.hide()
        elif selectiontype == "Institution":
            ui.institutionlongnrComboBox.show()
            ui.unitlongnrLineEdit.hide()
        elif selectiontype == "Unit":
            ui.institutionlongnrComboBox.hide()
            ui.unitlongnrLineEdit.show()

    querytupple = []

    queryheading = []

    def generateexcelforquery(self):


        if len(self.querytupple) < 1:
            self.showtooltip("Query Data Not Found")
            return;

        name = QtGui.QFileDialog.getSaveFileName(directory=r"C:\Users\{}\Documents".format(os.getlogin()),
                                                 caption="Save File",
                                                 filter="Excel (*.xlsx);;CSV (*.csv)")
        if not name:
            return

        self.queryheading = [i.strip() for i in self.queryheading]
        data = pd.DataFrame(self.querytupple, columns=self.queryheading)
        try:
            if name.endswith('.csv'):
                data.to_csv(name, index=False, float_format='string')
            elif name.endswith('.xlsx'):
                data.to_excel(name, float_format='string')
            self.showtooltip("Results Saved Successfully")
        except:
            self.showtooltip("Saving Failed")
        os.startfile(name)

    def backupdata(self):
        loc = QtGui.QFileDialog.getExistingDirectory(caption='Select the location to Backup All data',
                                                     directory=r'C:\users\{}'.format(os.getlogin()))

        if not loc:
            return


        try:
            ref = 0
            try:
                copy2('ncc.db', loc)
                ref = 1
                copy2('settings.ini', loc)
            except:
                if ref == 0:
                    QtGui.QMessageBox.critical(ui.Enrol, "Database backup Failed !",
                                               r'DataBase backup has failed. Please backup up the database file Manually. It is located in C:\ProgramFiles\NCC\ncc.db',
                                               'OK')

                elif ref == 1:
                    QtGui.QMessageBox.critical(ui.Enrol, "Settings.ini backup Failed !",
                                               r'Settings backup has failed. Please backup up the Settings.ini file Manually. It is located in C:\ProgramFiles\NCC\settings.ini',
                                               'OK')

            ref = 2
            if not os.path.exists(loc + r'\candidate photos'):
                os.mkdir(loc + r'\candidate photos')

            for i in os.listdir(r'candidate photos'):
                try:
                    copy2('candidate photos\\' + i, loc + '\candidate photos\\' + i)
                except:
                    pass;

        except Exception as e:
            if ref == 0:
                self.showtooltip("BACKUP FAILED")
                msg = "Database Settings and candidate photos BACKUP FAILED"
            if ref == 2:
                self.showtooltip("candidate photos BACKUP FAILED")
                msg = "Database BACKUP SUCCESSFULL\nCandidate photos BACKUP FAILED "
            QtGui.QMessageBox.warning(ui.Settings, 'Backup Problem',
                                      msg + "\nIf the problem persists Please backup the ncc.db file and candidate photos folder manually !")
            return

        self.showtooltip('BACKUP SUCCESSFULL')

    def restoredata(self):

        question = QtGui.QMessageBox()
        question.setWindowTitle('Choose the Type of file')
        question.setText("Do you want to Restore DATABASE FILE or Settings or CANDIDATES Photos ?")
        question.addButton('DataBase', QtGui.QMessageBox.YesRole)
        question.addButton('Settings', QtGui.QMessageBox.YesRole)
        question.addButton('Candidate Photos', QtGui.QMessageBox.YesRole)
        question.setStandardButtons(QtGui.QMessageBox.Cancel)

        ch = ''

        def onClicked(btn):
            nonlocal  ch
            ch = btn.text()

        question.buttonClicked.connect(onClicked)
        question.exec()


        # ch = QtGui.QMessageBox.question(ui.Settings, 'Choose the Type of file',
        #                                 'Do you want to Restore DATABASE FILE or Settings or CANDIDATES Photos ?', 'DataBase',
        #                                 "Candidate's Photos", "Settings",QtGui.QMessageBox.Cancel)
        if ch == 'DataBase':
            loc = QtGui.QFileDialog.getOpenFileNameAndFilter(caption='Select the Database file',
                                                             directory=r"C:\users\{}".format(os.getlogin()),
                                                             filter="Database (*.db)")
            if not loc[0]:
                return

            if QtGui.QMessageBox.question(ui.Settings, "Are you sure ? ",
                                          'Are you sure that you wish to restore the selected Database File ? Any DATA which is NOT in the selected file will be lost !',
                                          'Yes', 'No') == 0:

                try:
                    copy2(loc[0], os.getcwd() + r'\ncc.db')
                    self.showtooltip("Database Restored Successfully")
                except:
                    self.showtooltip("Database Restoration Failed")

        elif ch == 'Candidate Photos':
            loc = QtGui.QFileDialog.getExistingDirectory(caption='Select the Candidate Photos folder',
                                                         directory=r"C:\users\{}".format(os.getlogin()))
            if not loc:
                return
            try:
                if QtGui.QMessageBox.question(ui.Settings, 'Are you sure ? ',
                                              'Are you sure that you wish to restore the Photos in the Selected Folder ? The changes are irreversible !',
                                              'Yes', 'No') == 0:
                    for i in os.listdir(loc):
                        try:
                            copy2(loc + '\{}'.format(i), r'candidate photos\{}'.format(i))
                        except:
                            pass;
                else:
                    return
            except:
                self.showtooltip("RESTORATION FAILED")
                QtGui.QMessageBox.warning(ui.Settings, 'RESTORE FAILED',
                                          r'''Canidate Photos Restoration FAILED ! 
                                          Please copy the contents of your backedup photos to C:\Program Files\NCC\candidate photos Manually ! ''',
                                          'OK')
                return

            self.showtooltip("RESTORATION SUCCESSFULL")

        elif ch == 'Settings':
            loc = QtGui.QFileDialog.getOpenFileNameAndFilter(caption='Select the settings.ini file',
                                                             directory=r"C:\users\{}".format(os.getlogin()),
                                                             filter="Settings (*.ini)")
            if not loc[0]:
                return

            if QtGui.QMessageBox.question(ui.Settings, "Are you sure ? ",
                                          'Are you sure that you wish to restore the selected Settings.ini File ? Any DATA which is NOT in the selected file will be lost !',
                                          'Yes', 'No') == 0:

                try:
                    copy2(loc[0], os.getcwd() + r'\settings.ini')
                    self.showtooltip("Settings Restored Successfully")
                except:
                    self.showtooltip("Settings Restoration Failed")

    def save_from_excel_to_database(self):

        exceltodatabaseDialog = QtGui.QDialog()
        exceltodatabaseDialog.setObjectName(("exceltodatabaseDialog"))
        exceltodatabaseDialog.setModal(True)
        exceltodatabaseDialog.setWindowModality(QtCore.Qt.ApplicationModal)
        exceltodatabaseDialog.resize(778, 195)
        exceltodatabaseDialog.setMinimumSize(QtCore.QSize(778, 195))
        exceltodatabaseDialog.setMaximumSize(QtCore.QSize(778, 195))
        icon = QtGui.QIcon()
        icon.addPixmap(QtGui.QPixmap(_fromUtf8(":/icons/ncc2.png")), QtGui.QIcon.Normal, QtGui.QIcon.Off)
        exceltodatabaseDialog.setWindowIcon(icon)
        exceltodatabaseDialog.setStyleSheet(_fromUtf8("#exceltodatabaseDialog{\n"
                                                      "border-image:url(:/icons/simple_grad.png);\n"
                                                      "}"))
        verticalLayout = QtGui.QVBoxLayout(exceltodatabaseDialog)
        verticalLayout.setObjectName(_fromUtf8("verticalLayout"))
        label = QtGui.QLabel(exceltodatabaseDialog)
        font = QtGui.QFont()
        font.setFamily(_fromUtf8("Georgia"))
        font.setPointSize(20)
        label.setFont(font)
        label.setStyleSheet(_fromUtf8("color:white;"))
        label.setAlignment(QtCore.Qt.AlignCenter)
        label.setObjectName(_fromUtf8("label"))
        verticalLayout.addWidget(label)
        horizontalLayout = QtGui.QHBoxLayout()
        horizontalLayout.setObjectName(_fromUtf8("horizontalLayout"))
        etd_enrolmentPushButton = QtGui.QPushButton(exceltodatabaseDialog)
        font = QtGui.QFont()
        font.setFamily(_fromUtf8("Georgia"))
        font.setPointSize(13)
        font.setBold(False)
        font.setWeight(50)
        etd_enrolmentPushButton.setFont(font)
        etd_enrolmentPushButton.setObjectName(_fromUtf8("etd_enrolmentPushButton"))
        horizontalLayout.addWidget(etd_enrolmentPushButton)
        etd_acertPushButton = QtGui.QPushButton(exceltodatabaseDialog)
        font = QtGui.QFont()
        font.setFamily(_fromUtf8("Georgia"))
        font.setPointSize(13)
        font.setBold(False)
        font.setWeight(50)
        etd_acertPushButton.setFont(font)
        etd_acertPushButton.setObjectName(_fromUtf8("etd_acertPushButton"))
        horizontalLayout.addWidget(etd_acertPushButton)
        etd_bcertPushButton = QtGui.QPushButton(exceltodatabaseDialog)
        font = QtGui.QFont()
        font.setFamily(_fromUtf8("Georgia"))
        font.setPointSize(13)
        font.setBold(False)
        font.setWeight(50)
        etd_bcertPushButton.setFont(font)
        etd_bcertPushButton.setObjectName(_fromUtf8("etd_bcertPushButton"))
        horizontalLayout.addWidget(etd_bcertPushButton)
        etd_ccertPushButton = QtGui.QPushButton(exceltodatabaseDialog)
        font = QtGui.QFont()
        font.setFamily(_fromUtf8("Georgia"))
        font.setPointSize(13)
        font.setBold(False)
        font.setWeight(50)
        etd_ccertPushButton.setFont(font)
        etd_ccertPushButton.setObjectName(_fromUtf8("etd_ccertPushButton"))
        horizontalLayout.addWidget(etd_ccertPushButton)
        verticalLayout.addLayout(horizontalLayout)
        exceltodatabaseDialog.setWindowTitle(_translate("exceltodatabaseDialog", "Choose the TABLE", None))
        label.setText(
            _translate("exceltodatabaseDialog", "Select the Table to which you wish to add the Data", None))
        etd_enrolmentPushButton.setText(_translate("exceltodatabaseDialog", "Enrolment", None))
        etd_acertPushButton.setText(_translate("exceltodatabaseDialog", "A_Certificate", None))
        etd_bcertPushButton.setText(_translate("exceltodatabaseDialog", "B_Certificate", None))
        etd_ccertPushButton.setText(_translate("exceltodatabaseDialog", "C_Certificate", None))
        tab = ''

        def buttonpressed(obj):
            nonlocal tab
            tab = {'etd_acertPushButton': 'A_cert_marks', 'etd_bcertPushButton': 'B_cert_marks',
                   'etd_ccertPushButton': 'C_cert_marks', 'etd_enrolmentPushButton': 'enrolment'}[obj.objectName()]
            exceltodatabaseDialog.close()

        etd_enrolmentPushButton.clicked.connect(lambda: buttonpressed(etd_enrolmentPushButton))
        etd_acertPushButton.clicked.connect(lambda: buttonpressed(etd_acertPushButton))
        etd_bcertPushButton.clicked.connect(lambda: buttonpressed(etd_bcertPushButton))
        etd_ccertPushButton.clicked.connect(lambda: buttonpressed(etd_ccertPushButton))
        exceltodatabaseDialog.exec()

        if not tab: return

        con = connect(r'ncc.db')
        cur = con.cursor()

        loc = QtGui.QFileDialog.getOpenFileName(directory=r"C:\users\{}".format(os.getlogin()),
                                                caption="Select Excel or CSV file ",
                                                filter="Excel or CSV (*.xlsx *.csv)")

        if not loc:
            return
        if loc.endswith('.csv'):
            data = pd.read_csv(loc, na_filter=False)
        elif loc.endswith('.xlsx'):
            data = pd.read_excel(loc, na_filter=False)

        if tab != 'enrolment':
            if tab == 'A_cert_marks':
                data = data.iloc[3:, 0:24].reset_index(drop=True)
                data.columns = ['Enrolment_Number', 'Roll_Number', 'Rank', 'Student_Name',
                                'Fathers_Name', 'Date_Of_Birth', 'Enrol_Date', 'Camps_Attended',
                                'Date_Of_Discharge', 'Parade_Attendance_Year1',
                                'Parade_Attendance_Year2', 'Part1_Drill_Written',
                                'Part1_Drill_Practical', 'Part1_Drill_Total', 'Part2_WT_Written',
                                'Part2_WT_Practical', 'Part2_WT_Total', 'Part3_Misc_Written',
                                'Part4_SplSubjects_Written', 'Part4_SplSubjects_Practical',
                                'Part4_SplSubjects_Total', 'GrandTotal', 'Grading', 'Institution']

            elif tab == 'B_cert_marks':
                data = data.iloc[3:, 0:25].reset_index(drop=True)
                data.columns = ['Enrolment_Number', 'Roll_Number', 'Rank', 'Student_Name',
                                'Fathers_Name', 'Date_Of_Birth', 'Enrol_Date', 'Camps_Attended',
                                'Date_Of_Discharge', 'Parade_Attendance_Year1',
                                'Parade_Attendance_Year2', 'Part1_Drill_Written',
                                'Part1_Drill_Practical', 'Part1_Drill_Total', 'Part2_WT_Written',
                                'Part2_WT_Practical', 'Part2_WT_Total', 'Part3_Misc_Written',
                                'Part4_SplSubjects_Written', 'Part4_SplSubjects_Practical',
                                'Part4_SplSubjects_Total', 'Bonus_Marks_Certificate', 'GrandTotal',
                                'Grading', 'Institution']

            elif tab == 'C_cert_marks':
                data = data.iloc[3:, 0:25].reset_index(drop=True)
                data.columns = ['Enrolment_Number', 'Roll_Number', 'Rank', 'Student_Name',
                                'Fathers_Name', 'Date_Of_Birth', 'Enrol_Date', 'Camps_Attended',
                                'Date_Of_Discharge', 'Parade_Attendance_Year1',
                                'Parade_Attendance_Year2', 'Part1_Drill_Written',
                                'Part1_Drill_Practical', 'Part1_Drill_Total', 'Part2_WT_Written',
                                'Part2_WT_Practical', 'Part2_WT_Total', 'Part3_Misc_Written',
                                'Part4_SplSubjects_Written', 'Part4_SplSubjects_Practical',
                                'Part4_SplSubjects_Total', 'Bonus_Marks_Certificate', 'GrandTotal',
                                'Grading', 'Institution']

        else:
            data.columns = ['Enrolment_Number', 'Rank', 'Aadhaar_Number', 'Student_First_name', 'Student_Middle_Name', 'Student_Last_Name', 'Student_Name', 'Fathers_First_Name', 'Fathers_Middle_Name', 'Fathers_Last_Name', 'Fathers_Name', 'Mothers_First_Name', 'Mothers_Middle_Name', 'Mothers_Last_Name', 'Mothers_Name', 'Sex', 'Date_Of_Birth', 'Address', 'Email', 'Mobile_Number', 'Blood_Group', 'Nearest_Railway_Station', 'Nearest_Police_Station', 'Education', 'Education_Marks', 'Identification_mark', 'Criminal_Court', 'Name_of_School_College', 'Enroll_Permission', 'Earlier_candidate', 'Earlier_Enrolment_Number', 'Dismissed', 'Certificate', 'Camps_Attended', 'Extra_Curricular_Activities', 'Special_Achievements', 'Enrol_Date', 'Remarks', 'Meal_Preference', 'Bank_Name', 'Branch', 'Account_Name', 'Account_Number', 'IFSC_Code', 'MICR', 'Pan_Number', 'Institution', 'Unit','Seniority']

        conflicts = ''
        notexists = ''

        if tab == 'enrolment':
            try:
                data.to_sql(name=tab, con=con, if_exists='append', index=False)
                self.showtooltip("DATA SUCCESSFULLY ADDED TO DATABASE !!!")
                con.close()
                return;


            except:
                conflicts = ''
                for i in range(len(data.Enrolment_Number)):
                    try:
                        values = tuple(data.iloc[i].values)
                        cur.execute('''Insert into {} values {}'''.format(tab, values))
                    except(OperationalError, IntegrityError) as e:
                        conflicts += 'Enrolment_Number = {} : Aadhaar_Number = {}\n'.format(data.Enrolment_Number[i],
                                                                                            data.Aadhaar_Number[i])




        else:

            for i in range(len(data.Enrolment_Number)):
                try:
                    q = "SELECT EXISTS(SELECT Enrolment_Number FROM enrolment WHERE Enrolment_Number=\"{}\"".format(
                        data.Enrolment_Number[i]) + ")"
                    if not con.execute(q).fetchone()[0]:
                        notexists += 'Enrolment_Number = {}\n'.format(data.Enrolment_Number[i])

                    values = tuple(data.iloc[i].values)
                    cur.execute('''Insert into {} values {}'''.format(tab, values))
                except(OperationalError, IntegrityError) as e:
                    conflicts += 'Enrolment_Number = {}\n'.format(data.Enrolment_Number[i])

            cmsg = '''

The List of Enrolment Numbers which are NOT in the Enrolment Database Table are as shown below

{}
These Entries are added to the Certificates Table but Not to the Enrolment Table. Please add these Enrolment Numbers to the Enrolment Table if there is a need by submitting the numbers in the main Enrolment Form.
'''.format(notexists)
            ctitle = '''While adding the Data to the Certificate Database , its been found that some Enrolment Numbers are NOT PRESENT in the Enrolment Database Table. Please Add them to the Enroment Database if needed, in the Enrolment Tab</span></p></body></html>'''

        emsg = '''The List of Conflicts of either Enrolment_Number or Aadhaar_Number is as shown below :

{}
These entries are not added to the Database .\nIf you wish to update the database entries of present Students, Use the Data Entry tab or Update Entry checkbox in Enrolment Form. '''.format(
            conflicts)
        etitle = '''While adding the Data to the {} Database Table, some conflicts were found because of the already present Enrolment_Number or Aadhaar_Number present in the {} Table and hence these Enrolment_Numbers are not Added to the Database</span></p></body></html>'''.format(
            tab, tab)

        con.commit()
        con.close()

        if conflicts or notexists:

            if conflicts and notexists:
                title = \
                    '''{}
                    <p align="center" style=" margin-top:5px; margin-bottom:0px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;"><span style=" font-family:'century'; font-size:11pt;">Also {}</span></p></body></html>       
                    '''.format(etitle, ctitle)

                msg = emsg + '\n\n' + cmsg

            elif conflicts:
                title = etitle
                msg = emsg
            else:
                title = ctitle
                msg = cmsg

            Dialog = QtGui.QDialog()
            Dialog.setObjectName(_fromUtf8("Dialog"))
            Dialog.resize(804, 450)
            Dialog.setMaximumSize(QtCore.QSize(914, 16777215))
            font = QtGui.QFont()
            font.setFamily(_fromUtf8("Georgia"))
            Dialog.setFont(font)
            Dialog.setSizeGripEnabled(True)
            verticalLayout = QtGui.QVBoxLayout(Dialog)
            verticalLayout.setObjectName(_fromUtf8("verticalLayout"))
            textBrowser = QtGui.QTextBrowser(Dialog)
            textBrowser.setMaximumSize(QtCore.QSize(16777215, 125))
            font = QtGui.QFont()
            font.setFamily(_fromUtf8("century"))
            font.setPointSize(12)
            textBrowser.setFont(font)
            textBrowser.setObjectName(_fromUtf8("textBrowser"))
            verticalLayout.addWidget(textBrowser)
            gridLayout = QtGui.QGridLayout()
            gridLayout.setObjectName(_fromUtf8("gridLayout"))
            pushButton_2 = QtGui.QPushButton(Dialog)
            sizePolicy = QtGui.QSizePolicy(QtGui.QSizePolicy.Expanding, QtGui.QSizePolicy.Fixed)
            sizePolicy.setHorizontalStretch(0)
            sizePolicy.setVerticalStretch(0)
            sizePolicy.setHeightForWidth(pushButton_2.sizePolicy().hasHeightForWidth())
            pushButton_2.setSizePolicy(sizePolicy)
            font = QtGui.QFont()
            font.setPointSize(10)
            pushButton_2.setFont(font)
            pushButton_2.setObjectName(_fromUtf8("pushButton_2"))
            gridLayout.addWidget(pushButton_2, 0, 0, 1, 1, QtCore.Qt.AlignLeft)
            pushButton = QtGui.QPushButton(Dialog)
            font = QtGui.QFont()
            font.setPointSize(12)
            pushButton.setFont(font)
            pushButton.setObjectName(_fromUtf8("pushButton"))
            gridLayout.addWidget(pushButton, 0, 1, 1, 1, QtCore.Qt.AlignHCenter)
            pushButton_3 = QtGui.QPushButton(Dialog)
            pushButton_3.setText(_fromUtf8(""))
            pushButton_3.setFlat(True)
            pushButton_3.setObjectName(_fromUtf8("pushButton_3"))
            gridLayout.addWidget(pushButton_3, 0, 2, 1, 1)
            verticalLayout.addLayout(gridLayout)
            textBrowser_2 = QtGui.QTextBrowser(Dialog)
            font = QtGui.QFont()
            font.setFamily(_fromUtf8("Century"))
            font.setPointSize(12)
            textBrowser_2.setFont(font)
            textBrowser_2.setObjectName(_fromUtf8("textBrowser_2"))
            textBrowser_2.setText(msg)
            textBrowser_2.hide()
            verticalLayout.addWidget(textBrowser_2)

            QtCore.QObject.connect(pushButton_2, QtCore.SIGNAL(_fromUtf8("clicked()")), lambda: (
                textBrowser_2.show(), pushButton_2.setText('Hide Details')) if textBrowser_2.isHidden() else (
                textBrowser_2.hide(), pushButton_2.setText('Show Details')))
            QtCore.QObject.connect(pushButton, QtCore.SIGNAL(_fromUtf8("clicked()")), Dialog.close)
            QtCore.QMetaObject.connectSlotsByName(Dialog)
            Dialog.setWindowTitle(_translate("Dialog", "Conflicts Found !", None))
            textBrowser.setHtml(_translate("Dialog",
                                           "<!DOCTYPE HTML PUBLIC \"-//W3C//DTD HTML 4.0//EN\" \"http://www.w3.org/TR/REC-html40/strict.dtd\">\n"
                                           "<html><head><meta name=\"qrichtext\" content=\"1\" /><style type=\"text/css\">\n"
                                           "p, li { white-space: pre-wrap; }\n"
                                           "</style></head><body style=\" font-family:\'century\'; font-size:12pt; font-weight:400; font-style:normal;\">\n"
                                           "<p align=\"center\" style=\" margin-top:5px; margin-bottom:0px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;\"><span style=\" font-family:\'century\'; font-size:12pt;\">" + title + "",
                                           None))
            pushButton_2.setText(_translate("Dialog", "Show Details", None))
            pushButton.setText(_translate("Dialog", "OK", None))
            Dialog.exec()


    def set_camps_list(self):
        ui.settings_campslistListWidget.clear()
        ui.settings_campslistListWidget.setSpacing(1)
        ui.enrol_campsListWidget.clear()
        ui.campsNameuploaddataComboBox.clear()
        ui.settings_addcampsLineEdit.clear()

        ui.enrol_campsListWidget.setSpacing(1)
        for i in self.allcampslist:
            item = QtGui.QListWidgetItem()
            item.setText(i)
            item.setTextAlignment(QtCore.Qt.AlignHCenter | QtCore.Qt.AlignVCenter | QtCore.Qt.AlignCenter)
            font = QtGui.QFont()
            font.setFamily(_fromUtf8("caladea"))
            font.setPointSize(13)
            font.setBold(True)
            font.setWeight(75)
            item.setFont(font)
            brush = QtGui.QBrush(QtGui.QColor(72, 255, 52, 100))
            brush.setStyle(QtCore.Qt.Dense2Pattern)
            item.setBackground(brush)
            ui.enrol_campsListWidget.addItem(item)
            ui.campsNameuploaddataComboBox.addItem(item.text())
            ui.campsattendedqueryComboBox.addItem(item.text())

        for i in self.allcampslist:
            item = QtGui.QListWidgetItem()
            item.setText(i)
            item.setTextAlignment(QtCore.Qt.AlignHCenter | QtCore.Qt.AlignVCenter | QtCore.Qt.AlignCenter)
            font = QtGui.QFont()
            font.setFamily(_fromUtf8("Garamond"))
            font.setPointSize(15)
            font.setBold(True)
            font.setWeight(75)
            item.setFont(font)
            brush = QtGui.QBrush(QtGui.QColor(72, 255, 52, 100))
            brush.setStyle(QtCore.Qt.Dense2Pattern)
            item.setBackground(brush)
            ui.settings_campslistListWidget.addItem(item)

    def add_remove_camp(self, obj):

        '''ADD CAMP'''

        if obj.objectName() == 'settings_addcampPushButton':

            campname = ui.settings_addcampsLineEdit.displayText().strip()

            if not campname:
                self.showtooltip('Enter Camp name before clicking Add Camp')
                return

            if QtGui.QMessageBox.question(ui.Settings, "Are you sure ? ",
                                          'Are you sure that you want to add {} to the Camps List ?'.format(campname),
                                          'Yes', 'No') == 0:
                self.allcampslist.append(campname)
                self.settings.setValue('allcampslist', ',,,'.join(self.allcampslist))
                self.set_camps_list()


        elif obj.objectName() == 'settings_removecampPushButton':
            campselected = ui.settings_campslistListWidget.currentItem()
            campselected = '' if not campselected else campselected.text().strip()

            if not campselected:
                self.showtooltip("Select a Camp First")
                return

            if QtGui.QMessageBox.question(ui.Settings, 'Are you sure ?',
                                          'Are you sure that you want to Remove {} from the Camps Lisk ?'.format(
                                              campselected), 'Yes', 'No') == 0:
                self.allcampslist.remove(campselected)
                self.settings.setValue('allcampslist', ',,,'.join(self.allcampslist))
                ui.settings_removecampPushButton.hide()
                self.set_camps_list()

    def showtooltip(self, text):
        tt = QtGui.QToolTip
        myfont = QtGui.QFont()
        myfont.setFamily("caladea")
        myfont.setBold(True)
        myfont.setPointSize(20)
        tt.setFont(myfont)
        mywin = QtGui.QMainWindow.frameGeometry(MainWindow)
        pos = mywin.center()
        pos.setX(pos.x() - 6.5 * len(text));
        pos.setY(mywin.y())
        tt.showText(pos, text, MainWindow,
                    QtGui.QLineEdit.geometry(ui.settings_institutionListWidget))

    def settings_form_item_clicked(self):
        '''THis function is called whenever the user clicks on an item in the forms list of settings tab . It handles the show and hide of various elements and displays the corresponding
        field names in the field tab and also controls the add field combobox
        '''
        ui.settings_addfieldPushButton.show()
        ui.settings_removefieldPushButton.show()

        selected_text = ui.settings_formsListWidget.currentItem().text().strip()

        if selected_text in ['A certificate', 'B certificate', 'C certificate']:
            ui.settings_addfieldPushButton.hide()
            ui.settings_removefieldPushButton.hide()

        if selected_text not in ['Cadet details', 'Yoga day', 'Enrolment Nominal roll', 'Camp Nominal roll',
                                 'Scholarship NR', 'A certificate', 'B certificate', 'C certificate',
                                 'Speciman signature of cadets', 'TADA to cadets camps', 'TADA to cadets for exam']:
            ui.settings_removeformPushButton.show()
        else:
            ui.settings_removeformPushButton.hide()

        fieldslistsql = self.nametolistsql.get(selected_text)
        fieldslistsql = fieldslistsql if fieldslistsql else []

        fieldslistnotsql = self.nametolistnotsql.get(selected_text)
        fieldslistnotsql = fieldslistnotsql if fieldslistnotsql else []

        self.set_fields_list(fieldslistsql, fieldslistnotsql)

    def handler(self):
        settings = QtCore.QSettings('settings.ini', QtCore.QSettings.IniFormat)
        geo = MainWindow.geometry()
        settings.setValue('window_geometry',
                          ',,,'.join([str(geo.x()), str(geo.y()), str(geo.width()), str(geo.height())]))
        settings.setValue('current_tab', ui.mytab.currentIndex())

    def set_fields_list(self, sqllist, notsqllist):
        """This function is called after a new field is addded to the field list . This function sets the styles for the fields and puts them in the field list of the settings tab"""

        ui.settings_fieldsListWidget.clear()
        ui.settings_addfieldLineEdit.clear()
        ui.settings_fieldsListWidget.setSpacing(1)

        for ele in sqllist:
            item = QtGui.QListWidgetItem()
            item.setTextAlignment(QtCore.Qt.AlignHCenter | QtCore.Qt.AlignVCenter | QtCore.Qt.AlignCenter)
            item.setText(ele)
            font = QtGui.QFont()
            font.setFamily(_fromUtf8("Caladea"))
            font.setPointSize(14)
            font.setBold(False)
            font.setItalic(True)
            font.setWeight(63)
            item.setFont(font)
            item.setToolTip('Standard fields')
            brush = QtGui.QBrush(QtGui.QColor(120, 190, 255, 200))
            brush.setStyle(QtCore.Qt.Dense4Pattern)
            item.setBackground(brush)
            ui.settings_fieldsListWidget.addItem(item)

        for ele in notsqllist:
            if ele in sqllist:
                continue
            item = QtGui.QListWidgetItem()
            item.setTextAlignment(QtCore.Qt.AlignHCenter | QtCore.Qt.AlignVCenter | QtCore.Qt.AlignCenter)
            item.setText(ele)
            font = QtGui.QFont()
            item.setToolTip('Non standard fields')
            font.setFamily(_fromUtf8("Caladea"))
            font.setPointSize(14)
            font.setBold(False)
            font.setItalic(True)
            font.setWeight(62)
            item.setFont(font)
            brush = QtGui.QBrush(QtGui.QColor(170, 170, 255, 200))
            brush.setStyle(QtCore.Qt.Dense4Pattern)
            item.setBackground(brush)
            ui.settings_fieldsListWidget.addItem(item)

        ui.settings_fieldsknownRadioButton.hide()
        ui.settings_fieldsunknownRadioButton.hide()
        if ui.settings_formsListWidget.currentItem().text().strip() not in ['A certificate', 'B certificate',
                                                                            'C certificate']:
            ui.settings_removefieldPushButton.show()
        ui.settings_fieldsComboBox.hide()
        ui.settings_addfieldLineEdit.hide()
        # Set the elements that are not already in the corresponding field into the add field combobox

        ui.settings_fieldsComboBox.clear()

        for ele in self.all_enrolment_form_fieldslist:
            if ele not in sqllist:
                ui.settings_fieldsComboBox.addItem(ele)

    def set_forms_list(self):
        """Sets the style and label for the list items in formsListWidget of settings"""
        ui.settings_formsListWidget.clear()
        for inst in self.formslist:
            item = QtGui.QListWidgetItem()
            item.setText(inst)
            item.setTextAlignment(QtCore.Qt.AlignHCenter | QtCore.Qt.AlignVCenter | QtCore.Qt.AlignCenter)
            font = QtGui.QFont()
            font.setFamily(_fromUtf8("caladea"))
            font.setPointSize(15)
            font.setBold(True)
            font.setWeight(75)
            item.setFont(font)
            brush = QtGui.QBrush(QtGui.QColor(72, 255, 52, 100))
            brush.setStyle(QtCore.Qt.Dense2Pattern)
            item.setBackground(brush)
            ui.settings_formsListWidget.addItem(item)

        ui.settings_removeformPushButton.hide()

    def settings_form_field_add_remove(self, obj):
        '''Called when Add form or Add field buttons of the settings tab are clicked'''

        def removefield_button_states():
            ui.settings_fieldsknownRadioButton.hide()
            ui.settings_fieldsunknownRadioButton.hide()
            ui.settings_removefieldPushButton.show()
            ui.settings_fieldsComboBox.hide()
            ui.settings_addfieldLineEdit.hide()

        selectedform = ui.settings_formsListWidget.currentItem()
        selectedform = '' if not selectedform else selectedform.text().strip()
        selectedfield = ui.settings_fieldsListWidget.currentItem()
        selectedfield = '' if not selectedfield else selectedfield.text().strip()

        '''FORMS'''
        if obj.objectName() == 'settings_addformPushButton':

            if not ui.settings_addformLineEdit.displayText().strip():
                QtGui.QMessageBox.warning(ui.Settings, 'Entry field is blank',
                                          '\nEnter the name of the new form in the Entrybox that you wish to add and then Click "Add Form"',
                                          'OK')
                return

            if QtGui.QMessageBox.question(ui.Settings, 'Are you Sure ? ',
                                          'Are you sure that you wish to add a New Form ? This will add your form through out the software.',
                                          "Yes", "No") == 0:

                formname = ui.settings_addformLineEdit.displayText().strip()

                if formname in self.formslist:
                    QtGui.QMessageBox.warning(ui.Settings, 'Form already exists',
                                              'Make sure that the Entered form name is not already in the forms list and Enter a unique name for your Form and click "Add Form"',
                                              'OK')
                    return

                self.formslist.append(formname)
                self.settings.setValue('formslist', ',,,'.join(self.formslist))
                self.settings.setValue(formname.replace(' ', '_') + '_sql_fieldslist', '')
                self.showtooltip("Form Added")

                ui.formsComboBox.clear()
                ui.formsComboBox.addItems(self.formslist)

                ui.settings_addformLineEdit.clear()
                ui.settings_formsListWidget.clear()
                self.set_forms_list()
                return
            else:
                return

        if obj.objectName() == 'settings_removeformPushButton':
            if not selectedform:
                QtGui.QMessageBox.warning(ui.Settings, 'No Form Seleted',
                                          'Please select a form first before removing it.', 'OK')
                return

            if QtGui.QMessageBox.question(ui.Settings, 'Are you sure ?',
                                          'Are you sure that you want to DELETE the form "{}" from the forms list ? This will remove the form throughout the software'.format(
                                              selectedform), 'Yes', 'No') == 0:
                self.formslist.remove(selectedform)
                self.settings.setValue('formslist', ',,,'.join(self.formslist))
                self.showtooltip("Form Removed")
                try:
                    self.settings.remove(selectedform.replace(' ', '_') + "_sql_fieldslist")
                    self.settings.remove(selectedform.replace(' ', '_') + "_notsql_fieldslist")
                except:
                    pass;

                ui.formsComboBox.clear()
                ui.formsComboBox.addItems(self.formslist)

                self.set_forms_list()

        if obj.objectName() == 'settings_addfieldPushButton':
            ui.settings_removefieldPushButton.hide()

            if ui.settings_fieldsknownRadioButton.isHidden():
                def addfield_lineedit_combobox_decide():
                    if ui.settings_fieldsknownRadioButton.isChecked():
                        ui.settings_fieldsComboBox.show()
                        ui.settings_addfieldLineEdit.hide()
                        ui.settings_removefieldPushButton.hide()

                    else:
                        ui.settings_fieldsComboBox.hide()
                        ui.settings_addfieldLineEdit.show()

                ui.settings_fieldsknownRadioButton.toggled.connect(addfield_lineedit_combobox_decide)
                ui.settings_fieldsComboBox.show()
                ui.settings_fieldsknownRadioButton.show()
                ui.settings_fieldsunknownRadioButton.show()
                ui.settings_fieldsknownRadioButton.setChecked(True)


            else:
                if ui.settings_fieldsknownRadioButton.isChecked():
                    if QtGui.QMessageBox.question(ui.Settings, 'Are You Sure ? ',
                                                  'Are you sure that you want to add the field selected in the selection box to the field lists ? This will add the field through out the software.',
                                                  'Yes', 'No') == 0:

                        mylist = self.nametolistsql.get(selectedform)
                        mylist = [] if not mylist else mylist

                        mylist.append(ui.settings_fieldsComboBox.currentText())
                        self.settings.setValue(selectedform.replace(' ', '_') + '_sql_fieldslist',
                                               ',,,'.join(mylist))
                        self.nametolistsql.update({selectedform: mylist})
                        self.showtooltip("Field Added")

                        self.set_fields_list(
                            self.nametolistsql.get(selectedform) if self.nametolistsql.get(selectedform) else [],
                            self.nametolistnotsql.get(selectedform) if self.nametolistnotsql.get(selectedform) else [])
                        removefield_button_states()
                        return
                    else:
                        removefield_button_states()

                else:
                    if not ui.settings_addfieldLineEdit.displayText().strip():
                        QtGui.QMessageBox.warning(ui.Settings, 'Empty Field !',
                                                  'Make sure that the field entry box is not empty before seleting "Add field". Enter a field name in the Editing box provided and click "Add form" to add it to the fields list',
                                                  'OK')
                        removefield_button_states()
                        return

                    if QtGui.QMessageBox.question(ui.Settings, 'Are You Sure ? ',
                                                  'Are you sure that you want to add the entered field to the list of fields of the corresponding form ? This will add the field through out the software.',
                                                  'Yes', 'No') == 0:
                        mylist = self.nametolistnotsql.get(selectedform)
                        mylist = [] if not mylist else mylist
                        mylist.append(ui.settings_addfieldLineEdit.displayText().strip())
                        self.settings.setValue(selectedform.replace(' ', '_') + '_notsql_fieldslist',
                                               ',,,'.join(mylist))
                        self.nametolistnotsql.update({selectedform: mylist})
                        self.showtooltip("Field Added")
                        self.set_fields_list(
                            self.nametolistsql.get(selectedform) if self.nametolistsql.get(selectedform) else [],
                            self.nametolistnotsql.get(selectedform) if self.nametolistnotsql.get(selectedform) else [])
                        removefield_button_states()
                        return

        if obj.objectName() == 'settings_removefieldPushButton':
            if not selectedfield:
                QtGui.QMessageBox.warning(ui.Settings, 'No field seleted',
                                          'Please select a field first before removing it.', 'OK')
                return

            if QtGui.QMessageBox.question(ui.Settings, 'Are you sure ? ',
                                          'Are you sure that you wish to remove the field "{}" from the selected form ? It will remove the field through out the entire software'.format(
                                              selectedfield), 'Yes', 'No') == 0:

                sqllist = self.nametolistsql.get(selectedform)
                notsqllist = self.nametolistnotsql.get(selectedform)

                if selectedfield in sqllist:
                    self.nametolistsql.get(selectedform).remove(selectedfield);
                    self.settings.setValue(selectedform.replace(' ', '_') + '_sql_fieldslist',
                                           ',,,'.join(self.nametolistsql.get(selectedform)))
                    self.showtooltip("Field Removed")
                elif selectedfield in notsqllist:
                    self.nametolistnotsql.get(selectedform).remove(selectedfield);
                    self.settings.setValue(selectedform.replace(' ', '_') + '_notsql_fieldslist',
                                           ',,,'.join(self.nametolistnotsql.get(selectedform)))
                    self.showtooltip("Field Removed")

                self.set_fields_list(
                    self.nametolistsql.get(selectedform) if self.nametolistsql.get(selectedform) else [],
                    self.nametolistnotsql.get(selectedform) if self.nametolistnotsql.get(selectedform) else [])

    def institution_add_or_remove(self, button):
        """Called whenever user clicks on the add or remove button of the institution list in the settings tab"""

        if button.text().strip() == 'Add':
            if not ui.settings_instLineEdit.displayText().strip():
                QtGui.QMessageBox.warning(ui.Settings, "Empty field",
                                          '\n\nPlease enter an institution name and click "Add" to add it to the present list',
                                          "OK")
                return

            inst_name = ui.settings_instLineEdit.displayText().strip()

            self.institutionlist.append(inst_name)

            self.settings.setValue('institutionlist', ",,,".join(self.institutionlist))
            self.showtooltip("Institution Added")

            self.set_institutions_list()

        if button.text().strip() == 'Remove':

            selectedItem = ui.settings_institutionListWidget.currentItem()
            if not selectedItem:
                QtGui.QMessageBox.warning(ui.Settings, "Nothing Selected",
                                          '\n\nPlease select an institution first, to remove it.',
                                          "OK")
                return

            if QtGui.QMessageBox.question(ui.Settings, "Are you sure ?",
                                          '\n\nAre you sure you want to remove this Institution from the list of institutions ? \nThe changes are irreversible ! \nClick "Yes" to remove it.',
                                          'Yes', 'No') == 0:
                selected_text = selectedItem.text().strip()

                self.institutionlist.remove(selected_text)

                self.settings.setValue('institutionlist', ",,,".join(self.institutionlist))
                self.showtooltip("Institution Removed")

                self.set_institutions_list()
        ui.settings_instLineEdit.clear()

    def set_institutions_list(self):

        """
            Manages the hiding and showing of the buttons in the Institution list in the settings tab and also sets the institutions for all the comboboxes which have institutions
            This function is called whenever an item is added or removed to the institution list
        """
        ui.settings_institutionListWidget.clear()
        ui.settings_addinstitutionPushButton.show()
        ui.removeinstitutionPushButton.show()
        ui.settings_backinstPushButton.show()

        ui.settings_institutionListWidget.setSpacing(1)

        for inst in self.institutionlist:
            item = QtGui.QListWidgetItem()
            item.setText(inst)
            item.setTextAlignment(QtCore.Qt.AlignHCenter | QtCore.Qt.AlignVCenter | QtCore.Qt.AlignCenter)
            font = QtGui.QFont()
            font.setFamily(_fromUtf8("Garamond"))
            font.setPointSize(15)
            font.setBold(True)
            font.setWeight(75)
            item.setFont(font)
            brush = QtGui.QBrush(QtGui.QColor(72, 255, 52, 100))
            brush.setStyle(QtCore.Qt.Dense3Pattern)
            item.setBackground(brush)
            ui.settings_institutionListWidget.addItem(item)

        ui.institutionenrollComboBox.clear()
        ui.institutionenrollComboBox.addItems(self.institutionlist)

        ui.institutionqueryComboBox.clear()
        ui.institutionqueryComboBox.addItems(self.institutionlist)

        ui.institutionuploaddatacomboBox.clear()
        ui.institutionuploaddatacomboBox.addItems(self.institutionlist)

        ui.institutionlongnrComboBox.clear()
        ui.institutionlongnrComboBox.addItems(self.institutionlist)

        ui.settings_instLineEdit.hide()
        ui.settings_addPushButton.hide()
        ui.settings_backinstPushButton.hide()

    def saveuploadeddata(self):
        selectedInstitutionName = ui.institutionuploaddatacomboBox.currentText()
        selectedDataType = ui.typecomboBox.currentText()
        conn = connect('ncc.db')
        cur = conn.cursor()
        if selectedDataType == "Camps_Attended":
            enrolnumbers = ui.enrolmentuploaddataLineEdit.text().split(',')
            sqldata = ENROLMENT_FORM.enroll().execute(
                "select * from enrolment where institution='" + selectedInstitutionName + "'")
            sqldata1 = ENROLMENT_FORM.enroll().execute(
                "select * from camps_details where Institution='" + selectedInstitutionName +
                "' and camp_Attended='" + ui.campsNameuploaddataComboBox.currentText() + "'")
            msg = ""
            duplicate = ""
            for i in range(len(enrolnumbers)):
                flag = 0
                for j in range(len(sqldata)):
                    if enrolnumbers[i] == sqldata[j][0]:
                        flag = 1
                if flag == 0:
                    duplicate = duplicate + str(enrolnumbers[i]) + ","
                else:
                    flag = 0
                    for j in range(len(sqldata1)):
                        if enrolnumbers[i] != sqldata1[j][0]:
                            flag = 2
                        else:
                            flag = 0
                            break;
                    if flag != 2:
                        for k in range(len(sqldata1)):
                            for l in range(len(enrolnumbers)):
                                if sqldata1[k][0] == enrolnumbers[l] and sqldata1[k][
                                    1] == ui.campsNameuploaddataComboBox.currentText():
                                    flag = 1
                                    break
                            if flag == 1:
                                break
                    if flag == 0 or flag == 2:
                        sql = "insert into camps_details values('" + enrolnumbers[
                            i] + "','" + ui.campsNameuploaddataComboBox.currentText() + "','" + ui.locationLineEdit.text() + "','" + ui.startdateDateEdit.text() + "','" + ui.enddateDateEdit.text() + "','" + ui.institutionuploaddatacomboBox.currentText() + "')"
                        cur.execute(sql)
                        presentcamps = ENROLMENT_FORM.enroll().execute(
                            "select Camps_Attended from enrolment where Enrolment_Number='" + enrolnumbers[i] + "'")
                        if presentcamps[0][0].find(ui.campsNameuploaddataComboBox.currentText()):
                            if not len(presentcamps[0][0]):
                                presentcamps = str(presentcamps[0][0]  + ui.campsNameuploaddataComboBox.currentText())
                            else:
                                presentcamps = str(presentcamps[0][0] + "," + ui.campsNameuploaddataComboBox.currentText())
                            cur.execute("update enrolment set Camps_Attended='"+presentcamps+"' where Enrolment_Number='"+enrolnumbers[i]+"'")

                    else:
                        msg = msg + str(enrolnumbers[i]) + ","
            if len(msg) > 0:
                if QtGui.QMessageBox.warning(ui.Settings, 'Message',
                                             'Camp details of Enrolment Numbers : ' + msg + " already saved",
                                             'Update', 'Ignore') == 0:
                    eno = msg[0:-1].split(',')
                    for i in range(len(eno)):
                        cur.execute("delete from camps_details where Enrolment_Number='" + eno[
                            i] + "' and Camp_Attended='" + ui.campsNameuploaddataComboBox.currentText() + "'")

                        sql = "insert into camps_details values('" + eno[
                            i] + "','" + ui.campsNameuploaddataComboBox.currentText() + "','" + ui.locationLineEdit.text() + "','" + ui.startdateDateEdit.text() + "','" + ui.enddateDateEdit.text() + "','" + ui.institutionuploaddatacomboBox.currentText() + "')"
                        cur.execute(sql)

            if len(duplicate) > 0:
                QtGui.QMessageBox.warning(ui.Settings, 'Message',
                                          "Enrolment Numbers : " + duplicate + " does not exist in database\nRemaining data is saved",
                                          'OK')
            conn.commit()
            self.showtooltip("sucessfull")
        elif selectedDataType == "A certificate" or selectedDataType == "B certificate" or selectedDataType == "C certificate":
            fieldsListSql = self.nametolistsql.get(selectedDataType)
            fieldsListNotSql = self.nametolistnotsql.get(selectedDataType)
            if selectedDataType == "A certificate":
                selectedDataType = "A_cert_marks"
            if selectedDataType == "B certificate":
                selectedDataType = "B_cert_marks"
            if selectedDataType == "C certificate":
                selectedDataType = "C_cert_marks"
            sql = "insert or replace into " + selectedDataType + " values("
            for i in range(ui.tableWidget.rowCount()):
                if i != 0:
                    sql = sql + "),("
                for j in range(ui.tableWidget.columnCount()):

                    if ui.tableWidget.horizontalHeaderItem(j).text() == "Rank":
                        sql = sql + "'" + str(self.rankuploadcombobox[i].currentText()) + "'"
                    else:
                        sql = sql + "'" + str(ui.tableWidget.item(i, j).text()) + "'"
                    if j != len(fieldsListSql) - 1:
                        sql = sql + ","

            sql = sql + ")"
            ENROLMENT_FORM.enroll().insertionexecute(sql)
            self.showtooltip("Saved")

        elif selectedDataType == "Attendance":
            conn=connect("ncc.db")
            cur=conn.cursor()
            year = ui.yearComboBox.currentText().replace(' ', '_')
            certificate = ui.certificateComboBox.currentText().replace(' ', '_')
            s1 = ENROLMENT_FORM.enroll().execute(
                "select Enrolment_Number,certificate,year from attendance where institution='" + ui.institutionuploaddatacomboBox.currentText() + "'")
            s2 = ENROLMENT_FORM.enroll().execute(
                "select Enrolment_Number from enrolment where institution='" + ui.institutionuploaddatacomboBox.currentText() + "'")
            sql = "insert into Attendance(Enrolment_Number," + certificate + "_" + year + "_total_days," + certificate + "_" + year + \
                  "_present_days," + certificate + "_" + year + ",eligability,certificate,institution,year) values("
            for i in range(ui.tableWidget.rowCount()):
                if i != 0:
                    sql = sql + ",("
                length = len(s1)
                fl = 0
                if length < ui.tableWidget.columnCount():
                    length = ui.tableWidget.columnCount()
                    fl = 1
                for j in range(length):
                    if fl == 1:
                        if j < len(s1):
                            if s2[i][0] == s1[j][0] and ui.certificateComboBox.currentText()[0:1] == s1[j][
                                1] and year == s1[j][2]:
                                cur.execute("delete from Attendance where Enrolment_Number='" + s2[i][0] + "' and certificate='" + ui.certificateComboBox.currentText()[
                                                                                 0:1] + "' and year='"+year+"'")
                        sql = sql + "'" + self.lineeditattendance[i][j].text() + "',"
                    if fl == 0:
                        if s2[i][0] == s1[j][0] and ui.certificateComboBox.currentText()[0:1] == s1[j][1] and year == \
                                s1[j][2]:
                            cur.execute("delete from Attendance where Enrolment_Number='" + s2[i][0] + "' and certificate='" +
                                ui.certificateComboBox.currentText()[0:1] + "' and year='"+year+"'")
                        if j < ui.tableWidget.columnCount():
                            sql = sql + "'" + self.lineeditattendance[i][j].text() + "',"
                sql = sql + "'" + ui.certificateComboBox.currentText()[
                                  0:1] + "','" + ui.institutionuploaddatacomboBox.currentText() + "','" + year + "')"
            cur.execute(sql)
            self.showtooltip("Sucessfully Inserted")
            conn.commit()
            conn.close()


            """year = ui.yearComboBox.currentText().replace(' ', '_')
            certificate = ui.certificateComboBox.currentText().replace(' ', '_')
            s1 = ENROLMENT_FORM.enroll().execute(
                "select Enrolment_Number,certificate,year from attendance where institution='" + ui.institutionuploaddatacomboBox.currentText() + "'")
            s2 = ENROLMENT_FORM.enroll().execute(
                "select Enrolment_Number from enrolment where institution='" + ui.institutionuploaddatacomboBox.currentText() + "'")
            sql = "insert or replace into Attendance(Enrolment_Number," + certificate + "_" + year + "_total_days," + certificate + "_" + year + \
                  "_present_days," + certificate + "_" + year + ",eligability,certificate,institution,year) values("
            for i in range(ui.tableWidget.rowCount()):
                if i != 0:
                    sql = sql + ",("
                for j in range(ui.tableWidget.columnCount()):
                    sql = sql + "'" + self.lineeditattendance[i][j].text() + "',"
                sql = sql + "'" + ui.certificateComboBox.currentText()[
                                  0:1] + "','" + ui.institutionuploaddatacomboBox.currentText() + "','" + year + "')"
            ENROLMENT_FORM.enroll().insertionexecute(sql)

            self.showtooltip("Sucessfully Inserted")
"""







        else:
            sql = "select Enrolment_Number," + selectedDataType + " from enrolment where institution='" + selectedInstitutionName + "'"
            sqldata = ENROLMENT_FORM.enroll().execute(sql)
            for i in range(len(sqldata)):
                sql1 = "update enrolment set " + selectedDataType + "='" + ui.tableWidget.item(i,1).text() + "' where Enrolment_Number='" + \
                       sqldata[i][0] + "'"
                cur.execute(sql1)
            conn.commit()
            self.showtooltip("Sucessfully Inserted")
        self.openuploaddata()
    def saveexceluploadeddata(self):
        data = []
        name = QtGui.QFileDialog.getSaveFileName(directory=r"C:\Users\{}\Documents".format(os.getlogin()),
                                                 caption="Save File",
                                                 filter="Excel (*.xlsx)")
        horizontalheader = []
        if not len(name):
            return
        if ui.typecomboBox.currentText() == "A certificate":
            book = load_workbook('A_CERTIFICATES.xlsx')
            sheet = book.get_sheet_by_name('A')
        elif ui.typecomboBox.currentText() == "B certificate":
            book = load_workbook('B_CERTIFICATES.xlsx')
            sheet = book.get_sheet_by_name('B')
        elif ui.typecomboBox.currentText() == "C certificate":
            book = load_workbook('C_CERTIFICATES.xlsx')
            sheet = book.get_sheet_by_name('C')
        else:
            book = Workbook()
            sheet = book.active
            for i in range(ui.tableWidget.columnCount()):
                horizontalheader.append(ui.tableWidget.horizontalHeaderItem(i).text())
            sheet.append(horizontalheader)

        for i in range(ui.tableWidget.rowCount()):
            if not ui.tableWidget.isRowHidden(i):
                data.append([])
                for j in range(ui.tableWidget.columnCount()):
                    if ui.tableWidget.horizontalHeaderItem(j).text() == "Rank":
                        txt = self.rankuploadcombobox[i].currentText()
                    elif ui.tableWidget.item(i, j) != None:
                        txt = ui.tableWidget.item(i, j).text()
                    data[i].append(txt)
        if len(data) < 1:
            self.showtooltip("No data found")
        for row in data:
            sheet.append(row)
        book.save(name)
        book.save(TemporaryFile())
        self.showtooltip("Excel file created sucessfully")
        os.startfile(name)

    rankuploadcombobox = []

    campsattendedcombobox = []

    lineeditattendance = []

    def typecomboboxlogic(self):
        if ui.typecomboBox.currentText() == 'Camps_Attended':
            ui.eligibilityCheckBox.hide()
            ui.startdateDateEdit.show()

            ui.startdateLabel.show()

            ui.enddateDateEdit.show()

            ui.enddateLabel.show()

            ui.enrolmentuploaddataLineEdit.show()

            ui.locationLineEdit.show()

            ui.campsNameuploaddataComboBox.show()

            ui.save_data_excelPushButton.show()

            ui.certificateComboBox.hide()

            ui.yearComboBox.hide()

        elif ui.typecomboBox.currentText() == 'Extra_Curricular_Activities' or ui.typecomboBox.currentText() == 'Remarks':
            ui.eligibilityCheckBox.hide()
            ui.save_data_excelPushButton.hide()

            ui.startdateDateEdit.hide()

            ui.startdateLabel.hide()

            ui.enddateDateEdit.hide()

            ui.enddateLabel.hide()

            ui.enrolmentuploaddataLineEdit.hide()

            ui.locationLineEdit.hide()

            ui.campsNameuploaddataComboBox.hide()

            ui.yearComboBox.hide()

            ui.certificateComboBox.hide()
        elif ui.typecomboBox.currentText() == "Attendance":
            ui.eligibilityCheckBox.hide()

            ui.yearComboBox.show()

            ui.certificateComboBox.show()

            ui.save_data_excelPushButton.hide()

            ui.startdateDateEdit.hide()

            ui.startdateLabel.hide()

            ui.enddateDateEdit.hide()

            ui.enddateLabel.hide()

            ui.enrolmentuploaddataLineEdit.hide()

            ui.locationLineEdit.hide()

            ui.campsNameuploaddataComboBox.hide()
        else:
            ui.eligibilityCheckBox.show()
            ui.yearComboBox.hide()

            ui.certificateComboBox.hide()

            ui.save_data_excelPushButton.show()

            ui.startdateDateEdit.hide()

            ui.startdateLabel.hide()

            ui.enddateDateEdit.hide()

            ui.enddateLabel.hide()

            ui.enrolmentuploaddataLineEdit.hide()

            ui.locationLineEdit.hide()

            ui.campsNameuploaddataComboBox.hide()
        self.openuploaddata()
    def lineEditUploadDataLogic(self):
        for i in range(len(self.lineeditattendance)):
            if len(self.lineeditattendance[i][2].text()) and len(self.lineeditattendance[i][1].text()):
                a = int(str(self.lineeditattendance[i][2].text()))
                b = int(str(self.lineeditattendance[i][1].text()))
                self.lineeditattendance[i][3].setText(str(a / b * 100))
                if a / b * 100 > 75.00:
                    self.lineeditattendance[i][4].setText("Yes")
                else:
                    self.lineeditattendance[i][4].setText("No")
            if len(self.lineeditattendance[i][2].text()):
                a = int(str(self.lineeditattendance[i][2].text()))
                b = int(str(self.lineeditattendance[i][1].text()))
                if a > b:
                    self.lineeditattendance[i][2].setText(str(int(a / 10)))

    hiderow = []

    def eligibilitylogic(self):
        myfont = QtGui.QFont()
        myfont.setBold(True)
        myfont.setFamily("georgia")
        selectedDataType = ui.typecomboBox.currentText()
        if selectedDataType == "A certificate":
            certificate = "A_cert_attendance"
        if selectedDataType == "B certificate":
            certificate = "B_cert_attendance"
        if selectedDataType == "C certificate":
            certificate = "C_cert_attendance"
        if ui.eligibilityCheckBox.isChecked():
            self.hiderow = []
            for i in range(ui.tableWidget.rowCount()):
                enrolment = ui.tableWidget.verticalHeaderItem(i).text()
                sqlattendance = ENROLMENT_FORM.enroll().execute(
                    "select " + certificate + "_1_year from attendance where institution='" + ui.institutionuploaddatacomboBox.currentText() + "' and certificate='" + ui.typecomboBox.currentText()[
                                                                                                                                                                       0:1] + "' and Enrolment_number='" + enrolment + "' and year='1_year'")
                sqlattendance1 = ENROLMENT_FORM.enroll().execute(
                    "select " + certificate + "_2_year from attendance where institution='" + ui.institutionuploaddatacomboBox.currentText() + "' and certificate='" + ui.typecomboBox.currentText()[
                                                                                                                                                                       0:1] + "' and Enrolment_number='" + enrolment + "' and year='2_year'")
                if len(sqlattendance) > 0 and len(sqlattendance1) > 0:
                    if len(sqlattendance[0]) > 0 and len(sqlattendance1[0]) > 0:
                        if len(str(sqlattendance[0][0])) > 0 and len(str(sqlattendance1[0][0])) > 0:
                            if len(str(sqlattendance[0][0])):
                                if sqlattendance[0][0] > 75.00:
                                    ui.tableWidget.setItem(i, 9, QtGui.QTableWidgetItem(str(sqlattendance[0][0])))
                                    ui.tableWidget.item(i, 9).setBackground(QtGui.QColor(170, 170, 170, 80))
                                    ui.tableWidget.item(i, 9).setFont(myfont)
                                    ui.tableWidget.item(i, 9).setTextAlignment(QtCore.Qt.AlignCenter)
                                else:
                                    ui.tableWidget.hideRow(i)
                                    self.hiderow.append(i)
                            if len(str(sqlattendance1[0][0])):
                                if sqlattendance1[0][0] > 75.00:
                                    ui.tableWidget.setItem(i, 10, QtGui.QTableWidgetItem(str(sqlattendance1[0][0])))
                                    ui.tableWidget.item(i, 10).setBackground(QtGui.QColor(170, 170, 170, 80))
                                    ui.tableWidget.item(i, 10).setFont(myfont)
                                    ui.tableWidget.item(i, 10).setTextAlignment(QtCore.Qt.AlignCenter)
                                else:
                                    ui.tableWidget.hideRow(i)
                                    self.hiderow.append(i)
                        else:
                            ui.tableWidget.hideRow(i)
                            self.hiderow.append(i)
                    else:
                        ui.tableWidget.hideRow(i)
                        self.hiderow.append(i)
                else:
                    ui.tableWidget.hideRow(i)
                    self.hiderow.append(i)

        else:
            for i in range(len(self.hiderow)):
                ui.tableWidget.showRow(self.hiderow[i])

    def openuploaddata(self):
        if ui.typecomboBox.currentText()=="Select Type":
            ui.tableWidget.setStyleSheet("background-color:transparent;")
            ui.tableWidget.setRowCount(0)
            ui.tableWidget.setColumnCount(0)
            return
        else:
            ui.tableWidget.setStyleSheet("background-color:white;")
        ui.tableWidget.setEditTriggers(QtGui.QAbstractItemView.AllEditTriggers)
        ui.tableWidget.setRowCount(0)
        ui.tableWidget.setColumnCount(0)
        self.rankuploadcombobox = []
        self.rank = ["Cadet (CDT)", "Lance Corporal (LCPL)", "Corporal (CPL)", "Sergent (SGT)",
                     "Company Sergent Major (CSM)", "Junior Under Officer (JUO)", "Senior Under Officer (SUO)"]
        self.camps = ["NIC", "CATC", "AAC"]
        selectedInstitutionName = ui.institutionuploaddatacomboBox.currentText()
        selectedDataType = ui.typecomboBox.currentText()

        sql11 = "select Enrolment_Number from enrolment where institution='" + selectedInstitutionName + "'"
        verticalheaderdata = ENROLMENT_FORM.enroll().execute(sql11)
        verticalheader = []
        for i in range(len(verticalheaderdata)):
            verticalheader.append(verticalheaderdata[i][0])
        if selectedDataType == "Camps_Attended":
            verticalheader = []
            ui.tableWidget.clearContents()
            sql = "select Enrolment_Number from enrolment where institution='" + selectedInstitutionName + "'"
            sqldata = ENROLMENT_FORM.enroll().execute(sql)
            sqldata1 = ENROLMENT_FORM.enroll().execute(
                "select * from camps_details where Institution='" + selectedInstitutionName + "' and camp_Attended='" + ui.campsNameuploaddataComboBox.currentText() + "'")
            header = ["Enrolment_Number", "Camp_Attended", "Location", "Started_Date", "Ended_Date"]
            ui.tableWidget.setColumnCount(len(header))
            ui.tableWidget.setRowCount(len(sqldata1))
            ui.tableWidget.setHorizontalHeaderLabels(header)
            for i in range(ui.tableWidget.rowCount()):
                verticalheader.append(sqldata1[i][0])
                for j in range(ui.tableWidget.columnCount()):
                    if len(sqldata1[i]) > 0:
                        ui.tableWidget.setItem(i, j, QtGui.QTableWidgetItem(sqldata1[i][j]))

                    else:
                        ui.tableWidget.setItem(i, j, QtGui.QTableWidgetItem(""))
            ui.tableWidget.setVerticalHeaderLabels(verticalheader)
            ui.tableWidget.setEditTriggers(QtGui.QAbstractItemView.NoEditTriggers)
        elif selectedDataType == "A certificate" or selectedDataType == "B certificate" or selectedDataType == "C certificate":
            ui.eligibilityCheckBox.setChecked(False)
            certificate = ""
            ui.tableWidget.clearContents()
            fieldsListSql = self.nametolistsql.get(selectedDataType)
            if selectedDataType == "A certificate":
                selectedDataType = "A_cert_marks"
                certificate = "A_cert_attendance"
            if selectedDataType == "B certificate":
                selectedDataType = "B_cert_marks"
                certificate = "B_cert_attendance"
            if selectedDataType == "C certificate":
                selectedDataType = "C_cert_marks"
                certificate = "C_cert_attendance"
            sql = """select Enrolment_Number,Rank,Student_Name,Fathers_Name,Date_Of_Birth,Enrol_Date,Camps_Attended from enrolment where institution='""" + selectedInstitutionName + "'"
            sqldata = ENROLMENT_FORM.enroll().execute(sql)
            field = ""
            field = field[0:-1]
            sqlpresentdata = ENROLMENT_FORM.enroll().execute(
                "select * from " + selectedDataType + " where Institution='" + selectedInstitutionName + "'")
            ui.tableWidget.setRowCount(len(sqldata))
            ui.tableWidget.setColumnCount(len(fieldsListSql))
            ui.tableWidget.setHorizontalHeaderLabels(fieldsListSql)
            ui.tableWidget.setVerticalHeaderLabels(verticalheader)
            l = 0
            for i in range(len(sqldata)):
                flag = 0
                con = 0
                for l in range(len(sqlpresentdata)):
                    if sqldata[i][0] == sqlpresentdata[l][0]:
                        flag = 1
                        break
                if flag == 0:
                    for j in range(len(fieldsListSql)):
                        if j < len(sqldata[i]):
                            if fieldsListSql[j] == "Roll_Number":
                                ui.tableWidget.setItem(i, j, QtGui.QTableWidgetItem(""))
                                con = 1
                            if con == 0:
                                if fieldsListSql[j] == "Rank":
                                    self.rankuploadcombobox.append(QComboBox(ui.tableWidget))
                                    ui.tableWidget.setCellWidget(i, j, self.rankuploadcombobox[
                                        len(self.rankuploadcombobox) - 1])
                                    for items in range(len(self.rank)):
                                        self.rankuploadcombobox[len(self.rankuploadcombobox) - 1].addItem(_fromUtf8(""))
                                        self.rankuploadcombobox[len(self.rankuploadcombobox) - 1].setItemText(items,
                                                                                                              _translate(
                                                                                                                  "MainWindow",
                                                                                                                  self.rank[
                                                                                                                      items],
                                                                                                                  None))
                                    self.rankuploadcombobox[len(self.rankuploadcombobox) - 1].setStyleSheet(
                                        "font-weight:bold;")
                                    self.rankuploadcombobox[len(self.rankuploadcombobox) - 1].setCurrentIndex(
                                        self.rankuploadcombobox[len(self.rankuploadcombobox) - 1].findText(
                                            sqldata[i][j]))
                                else:
                                    ui.tableWidget.setItem(i, j, QtGui.QTableWidgetItem(sqldata[i][j]))
                            if con == 1:
                                if fieldsListSql[j] == "Roll_Number":
                                    self.rankuploadcombobox.append(QComboBox(ui.tableWidget))
                                    ui.tableWidget.setCellWidget(i, j + 1, self.rankuploadcombobox[
                                        len(self.rankuploadcombobox) - 1])
                                    for items in range(len(self.rank)):
                                        self.rankuploadcombobox[len(self.rankuploadcombobox) - 1].addItem(_fromUtf8(""))
                                        self.rankuploadcombobox[len(self.rankuploadcombobox) - 1].setItemText(items,
                                                                                                              _translate(
                                                                                                                  "MainWindow",
                                                                                                                  self.rank[
                                                                                                                      items],
                                                                                                                  None))
                                    self.rankuploadcombobox[len(self.rankuploadcombobox) - 1].setStyleSheet(
                                        "font-weight:bold;")
                                    self.rankuploadcombobox[len(self.rankuploadcombobox) - 1].setCurrentIndex(
                                        self.rankuploadcombobox[len(self.rankuploadcombobox) - 1].findText(
                                            sqldata[i][j]))
                                else:
                                    ui.tableWidget.setItem(i, j + 1, QtGui.QTableWidgetItem(sqldata[i][j]))
                        else:
                            if j != len(fieldsListSql) - 2:
                                ui.tableWidget.setItem(i, j + 1, QtGui.QTableWidgetItem(""))
                            if j == len(fieldsListSql) - 2:
                                ui.tableWidget.setItem(i, j + 1, QtGui.QTableWidgetItem(selectedInstitutionName))

                if flag == 1:
                    for j in range(len(sqlpresentdata[l])):
                        if fieldsListSql[j] == "Rank":
                            self.rankuploadcombobox.append(QComboBox(ui.tableWidget))
                            ui.tableWidget.setCellWidget(i, j,
                                                         self.rankuploadcombobox[len(self.rankuploadcombobox) - 1])
                            for items in range(len(self.rank)):
                                self.rankuploadcombobox[len(self.rankuploadcombobox) - 1].addItem(_fromUtf8(""))
                                self.rankuploadcombobox[len(self.rankuploadcombobox) - 1].setItemText(items, _translate(
                                    "MainWindow", self.rank[items], None))
                            self.rankuploadcombobox[len(self.rankuploadcombobox) - 1].setStyleSheet("font-weight:bold;")
                            self.rankuploadcombobox[len(self.rankuploadcombobox) - 1].setCurrentIndex(
                                self.rankuploadcombobox[len(self.rankuploadcombobox) - 1].findText(
                                    sqlpresentdata[l][j]))
                        else:
                            ui.tableWidget.setItem(i, j, QtGui.QTableWidgetItem(sqlpresentdata[l][j]))
                if len(sqlpresentdata) > 0:
                    sqlpresentdata.pop(l)
            for i in range(ui.tableWidget.rowCount()):
                enrolment = ui.tableWidget.verticalHeaderItem(i).text()
                sqlattendance = ENROLMENT_FORM.enroll().execute(
                    "select " + certificate + "_1_year from attendance where institution='" + ui.institutionuploaddatacomboBox.currentText() + "' and certificate='" + ui.typecomboBox.currentText()[
                                                                                                                                                                       0:1] + "' and Enrolment_number='" + enrolment + "' and year='1_year'")
                sqlattendance1 = ENROLMENT_FORM.enroll().execute(
                    "select " + certificate + "_2_year from attendance where institution='" + ui.institutionuploaddatacomboBox.currentText() + "' and certificate='" + ui.typecomboBox.currentText()[
                                                                                                                                                                       0:1] + "' and Enrolment_number='" + enrolment + "' and year='2_year'")

                if len(sqlattendance) > 0:
                    if len(str(sqlattendance[0][0])):
                        ui.tableWidget.setItem(i, 9, QtGui.QTableWidgetItem(str(sqlattendance[0][0])))
                if len(sqlattendance1) > 0:
                    if len(str(sqlattendance1[0][0])):
                        ui.tableWidget.setItem(i, 10, QtGui.QTableWidgetItem(str(sqlattendance1[0][0])))

        elif selectedDataType == "Attendance":
            ui.tableWidget.clearContents()
            sql = "select Enrolment_Number from enrolment where institution='" + ui.institutionuploaddatacomboBox.currentText() + "'"
            sqldata = ENROLMENT_FORM.enroll().execute(sql)
            sql1 = "select * from Attendance where institution='" + ui.institutionuploaddatacomboBox.currentText() + "' and certificate='" + ui.certificateComboBox.currentText()[
                                                                                                                                             0:1] + "'"
            sqldata1 = ENROLMENT_FORM.enroll().execute(sql1)
            attendancelist = ["Enrolment_Number", "certificate", "institution", "A_cert_attendance_1_year",
                              "A_cert_attendance_2_year", "B_cert_attendance_1_year", "B_cert_attendance_2_year",
                              "C_cert_attendance_1_year", "C_cert_attendance_2_year",
                              "A_cert_attendance_1_year_total_days", "A_cert_attendance_2_year_total_days",
                              "B_cert_attendance_1_year_total_days", 'B_cert_attendance_2_year_total_days',
                              "C_cert_attendance_1_year_total_days", "C_cert_attendance_2_year_total_days",
                              "A_cert_attendance_1_year_present_days", "A_cert_attendance_2_year_present_days",
                              "B_cert_attendance_1_year_present_days", "B_cert_attendance_2_year_present_days",
                              "C_cert_attendance_1_year_present_days", "C_cert_attendance_2_year_present_days"]

            horizontalheader = ['Enrolment_Number', 'Total Days', 'Attended Days', 'Percentage', 'Eligibility']
            ui.tableWidget.setRowCount(len(sqldata))
            ui.tableWidget.setColumnCount(len(horizontalheader))
            ui.tableWidget.setHorizontalHeaderLabels(horizontalheader)
            verticalheader = []
            self.lineeditattendance = []
            year = ui.yearComboBox.currentText().replace(' ', '_')
            certificate = ui.certificateComboBox.currentText().replace(' ', '_')
            for i in range(ui.tableWidget.rowCount()):
                flag = 0
                sqldata2 = []
                self.lineeditattendance.append([])
                verticalheader.append(sqldata[i][0])
                for l in range(len(sqldata1)):
                    sql = "select Enrolment_Number," + certificate + "_" + year + "_total_days," + certificate + "_" + year + "_present_days," + certificate + "_" + year + ",eligability from Attendance" \
                                                                                                                                                                            " where Enrolment_Number='" + \
                          sqldata[i][0] + "'and certificate='" + ui.certificateComboBox.currentText()[
                                                                 0:1] + "' and year='" + year + "'"
                    sqldata2 = ENROLMENT_FORM.enroll().execute(sql)
                    if len(sqldata2):
                        flag = 1
                        break
                if flag == 1:
                    for j in range(ui.tableWidget.columnCount()):
                        self.lineeditattendance[i].append(QtGui.QLineEdit(ui.tableWidget))
                        ui.tableWidget.setCellWidget(i, j, self.lineeditattendance[i][j])
                        if len(str(sqldata2[0][j])):
                            self.lineeditattendance[i][j].setText(_translate("MainWindow", str(sqldata2[0][j]), None))
                        else:
                            self.lineeditattendance[i][j].setText(_translate("MainWindow", "", None))
                        if j != 0 and j != 4:
                            self.lineeditattendance[i][j].setValidator(QtGui.QDoubleValidator())
                        if j == 1 or j == 2:
                            self.lineeditattendance[i][j].setMaxLength(2)
                            self.lineeditattendance[i][j].textChanged.connect(self.lineEditUploadDataLogic)
                        if j == 3:
                            self.lineeditattendance[i][j].setMaxLength(5)
                            self.lineeditattendance[i][j].setDisabled(True)
                        if j == 0:
                            self.lineeditattendance[i][j].setDisabled(True)
                        if j == 4:
                            self.lineeditattendance[i][j].setMaxLength(3)
                            self.lineeditattendance[i][j].setDisabled(True)
                        self.lineeditattendance[i][j].setStyleSheet("background-color:lightgray;")
                else:
                    for j in range(ui.tableWidget.columnCount()):
                        self.lineeditattendance[i].append(QtGui.QLineEdit(ui.tableWidget))
                        ui.tableWidget.setCellWidget(i, j, self.lineeditattendance[i][j])
                        if j == 0:
                            self.lineeditattendance[i][j].setText(_translate("MainWindow", sqldata[i][0], None))
                            self.lineeditattendance[i][j].setDisabled(True)
                        else:
                            self.lineeditattendance[i][j].setText(_translate("MainWindow", "", None))
                        if j != 0 and j != 4:
                            self.lineeditattendance[i][j].setValidator(QtGui.QDoubleValidator())
                        if j == 1 or j == 2:
                            self.lineeditattendance[i][j].setMaxLength(2)
                            self.lineeditattendance[i][j].textChanged.connect(self.lineEditUploadDataLogic)
                        if j == 3:
                            self.lineeditattendance[i][j].setMaxLength(5)
                            self.lineeditattendance[i][j].setDisabled(True)
                        if j == 4:
                            self.lineeditattendance[i][j].setMaxLength(3)
                            self.lineeditattendance[i][j].setDisabled(True)
                        self.lineeditattendance[i][j].setStyleSheet("background-color:lightgray;")
            ui.tableWidget.setVerticalHeaderLabels(verticalheader)
            ui.tableWidget.setStyleSheet("background-color:transparent;")
            ui.tableWidget.hideColumn(0)
        else:
            ui.tableWidget.setColumnCount(0)
            ui.tableWidget.setRowCount(0)
            sql = "select Enrolment_Number," + selectedDataType + " from enrolment where institution='" + selectedInstitutionName + "'"
            sqldata = ENROLMENT_FORM.enroll().execute(sql)
            ui.tableWidget.setColumnCount(2)
            ui.tableWidget.setRowCount(len(sqldata))
            header = ["Enrolment Number"]
            header.append(selectedDataType)
            ui.tableWidget.setHorizontalHeaderLabels(header)
            ui.tableWidget.setVerticalHeaderLabels(verticalheader)
            for i in range(len(sqldata)):
                for j in range(len(sqldata[i])):
                    ui.tableWidget.setItem(i, j, QtGui.QTableWidgetItem(sqldata[i][j]))
        myfont = QtGui.QFont()
        myfont.setBold(True)
        myfont.setFamily("georgia")
        for i in range(ui.tableWidget.rowCount()):
            for j in range(ui.tableWidget.columnCount()):
                if ui.tableWidget.item(i, j) != None:
                    ui.tableWidget.item(i, j).setBackground(QtGui.QColor(170, 170, 170, 80))
                    ui.tableWidget.item(i, j).setFont(myfont)
                    ui.tableWidget.item(i, j).setTextAlignment(QtCore.Qt.AlignCenter)
        ui.tableWidget.cellChanged.connect(self.textchanged)

        ui.tableWidget.showGrid()
        ui.tableWidget.setStyleSheet(
            "color:black;font-weight:bold;font-size:15px;border:1px solid black;gridline-color:black;")
        ui.tableWidget.horizontalHeader().setStyleSheet(
            "color:darkgreen;font-size:24px;font-weight:bold;font-family:'gabriola';border:1px solid black;")
        ui.tableWidget.verticalHeader().setStyleSheet(
            "color:darkorange;font-size:20px;font-weight:bold;font-family:caladea;border:1px solid black;gridline-color:black;")
        ui.tableWidget.resizeRowsToContents()
        ui.tableWidget.resizeColumnsToContents()
    def textchanged(self,i,j):
        ui.tableWidget.resizeColumnToContents(j)

    def conditionscomboboxlogic(self):

        text = ui.conditionlistcombobox.currentText()
        ui.rankqueryComboBox.hide()
        ui.institutionqueryComboBox.hide()
        ui.bloodgroupqueryComboBox.hide()
        ui.sexqueryComboBox.hide()
        ui.datequeryDateEdit.hide()
        ui.valuelineEdit.hide()
        ui.certificatequeryComboBox.hide()
        ui.vegitarianqueryComboBox.hide()
        ui.campsattendedqueryComboBox.hide()
        ui.seniorityqueryComboBox.hide()
        if text == "Rank":
            ui.rankqueryComboBox.show()
        elif text == "Institution":
            ui.institutionqueryComboBox.show()
        elif text == "Blood_Group":
            ui.bloodgroupqueryComboBox.show()
        elif text == "Sex":
            ui.sexqueryComboBox.show()
        elif text == "Enrol_Date" or text == "Date_Of_Birth":
            ui.datequeryDateEdit.show()
        elif text == "Meal_Preference":
            ui.vegitarianqueryComboBox.show()
        elif text == "Certificate":
            ui.certificatequeryComboBox.show()
        elif text == "Camps_Attended":
            ui.campsattendedqueryComboBox.show()
        elif text == "Seniority":
            ui.seniorityqueryComboBox.show()
        else:
            ui.valuelineEdit.show()

    def enrol_button_pressed(self):

        ui.searchbyfieldLineEdit.clear()
        self.enable_query_checkbox_elements();

    def update_excel_function(self):
        if ui.entryBox.toPlainText() == "":
            self.showtooltip('Please Enter the Enrolment numbers in the EntryBox.')
            return
        x = ui.entryBox.toPlainText().replace(' ', '')
        if not len(x):
            self.showtooltip("Enrolment Entry Box is empty")
            return ;

        enrolnocopy = x.split(',')
        enrolno = list(enrolnocopy)
        con = connect("ncc.db")
        cur = con.cursor()
        duplicate=""
        # enrollist = ENROLMENT_FORM.enroll().execute("select Enrolment_Number from enrolment")
        for i in enrolnocopy:
            if not cur.execute(
                    "select Exists(select Enrolment_Number from enrolment where Enrolment_Number='{}' Limit 1)".format(
                        i)).fetchone()[0]:
                enrolno.remove(i)
                duplicate = duplicate + str(i) + ","
        con.close()

        if not len(enrolno):
            self.showtooltip("Enrolment numbers are not enrolled.")
        selectedformname = ui.formsComboBox.currentText()
        self.listdata = self.nametolistsql.get(selectedformname)

        self.listheadingdata = list(self.nametolistsql.get(selectedformname))
        for i in self.nametolistnotsql.get(selectedformname):
            self.listheadingdata.append(i)

        sql = """select """
        if selectedformname != 'A certificate' and selectedformname != "B certificate" and selectedformname != "C certificate":
            for i in range(len(self.listdata)):
                sql = sql + self.listdata[i]
                if i != len(self.listdata) - 1:
                    sql = sql + ","
            sql = sql + " from enrolment where "
        else:
            tablename = ""
            if selectedformname == "A certificate":
                tablename = "A_cert_marks"
            elif selectedformname == "B certificate":
                tablename = "B_cert_marks"
            elif selectedformname == "C certificate":
                tablename = "C_cert_marks"
            for i in range(len(self.listdata)):
                sql = sql + self.listdata[i]
                if i != len(self.listdata) - 1:
                    sql = sql + ","
            sql = sql + " from " + tablename + " where "

        for i in range(len(enrolno)):
            sql = sql + " Enrolment_Number=\"" + enrolno[i] + "\" "
            if i != len(enrolno) - 1:
                sql = sql + "or"
        tup = ENROLMENT_FORM.enroll().execute(sql)
        if len(tup) == 0:
            QtGui.QMessageBox.warning(ui.Enrol, 'Message',
                                      'First Enter the Enrolment Numbers for the respective forms\nand then generate a form.',
                                      'OK')
            return
        self.formname = ""
        self.formname = QtGui.QFileDialog.getOpenFileName(directory=r"C:\Users\{}\Documents".format(os.getlogin()),
                                                          caption="Select the Form file",
                                                          filter="Excel (*.xlsx);;CSV (*.csv)")
        if self.formname == "":
            return

        if tup[0] != len(self.listheadingdata):
            for i in range(len(tup)):
                cp = list(tup[i])
                while (len(cp) != len(self.listheadingdata)):
                    cp.append('')
                tup[i] = cp

        if self.formname.endswith(".csv"):
            pd.DataFrame(tup, columns=self.listheadingdata).to_csv(self.formname, mode='a', index=False, header=False)
        elif self.formname.endswith('.xlsx'):
            data = pd.DataFrame(tup, columns=self.listheadingdata)
            data1 = pd.read_excel(self.formname, na_filter=False)
            finaldata = data1.append(data, ignore_index=True)
            finaldata.to_excel(self.formname)

        self.table1(tup, sql)
        if len(duplicate) == 0:
            self.showtooltip("Form updated sucessfully")
        else:
            self.showtooltip("Form generated sucessfully")
            QtGui.QMessageBox.warning(ui.enrolformFrame, "Message",
                                      "Form generated sucessfully\n Enrolment numbers : " + duplicate + " are not enrolled.",
                                      'OK')
        ui.entryBox.setText("")
        os.startfile(self.formname)

    def saveExcelfuntion(self):
        # formlist=["Cadet details","Yoga day","Enrolment Nominal roll","Camp Nominal roll","Scholarship NR","A certificate","B certificate","C certificate","Speciman signature of cadets","TADA to cadets camps","TADA to cadets for exam"]
        x = ui.entryBox.toPlainText().replace(" ", "")
        if not x.strip():
            self.showtooltip("Please Enter the Enrolment numbers in the EntryBox.")
            return

        x = x.replace("\n", ",")
        enrolnocopy = x.split(',')
        enrolno = list(enrolnocopy)
        con = connect("ncc.db")
        cur = con.cursor()
        duplicate = ""
        # enrollist = ENROLMENT_FORM.enroll().execute("select Enrolment_Number from enrolment")
        for i in enrolnocopy:
            if not cur.execute(
                    "select Exists(select Enrolment_Number from enrolment where Enrolment_Number='{}' Limit 1)".format(
                        i)).fetchone()[0]:
                enrolno.remove(i)
                duplicate = duplicate + str(i) + ","
        con.close()

        selectedformname = ui.formsComboBox.currentText()
        self.listdata = self.nametolistsql.get(selectedformname)
        self.listheadingdata = list(self.nametolistsql.get(selectedformname))
        for i in self.nametolistnotsql.get(selectedformname):
            self.listheadingdata.append(i)

        sql = """select """
        if selectedformname != 'A certificate' and selectedformname != "B certificate" and selectedformname != "C certificate":
            for i in range(len(self.listdata)):
                sql = sql + self.listdata[i]
                if i != len(self.listdata) - 1:
                    sql = sql + ","
            sql = sql + " from enrolment where "
        else:
            tablename = ""
            if selectedformname == "A certificate":
                tablename = "A_cert_marks"
            elif selectedformname == "B certificate":
                tablename = "B_cert_marks"
            elif selectedformname == "C certificate":
                tablename = "C_cert_marks"
            for i in range(len(self.listdata)):
                sql = sql + self.listdata[i]
                if i != len(self.listdata) - 1:
                    sql = sql + ","
            sql = sql + " from " + tablename + " where "
        if not len(enrolno):
            self.showtooltip('Make sure that entered Enrolment numbers are enrolled.')
            return
        for i in range(len(enrolno)):
            sql = sql + " Enrolment_Number=\"" + enrolno[i] + "\" "
            if i != len(enrolno) - 1:
                sql = sql + "or"
        tup = ENROLMENT_FORM.enroll().execute(sql)
        if not len(tup):
            self.showtooltip('Make sure that you have entered some data in the Data Entry table .')
            return
        self.formname = ""
        self.formname = QtGui.QFileDialog.getSaveFileName(directory=r"C:\Users\{}\Documents".format(os.getlogin()),
                                                          caption="Save File",
                                                          filter="Excel (*.xlsx);;CSV (*.csv)")
        if self.formname == "":
            return

            # kar1,kar2,kar3,kar4

        if tup[0] != len(self.listheadingdata):
            for i in range(len(tup)):
                cp = list(tup[i])
                while (len(cp) != len(self.listheadingdata)):
                    cp.append('')
                tup[i] = cp

        if self.formname.endswith(".csv"):
            pd.DataFrame(tup, columns=self.listheadingdata).to_csv(self.formname, index=False)
        elif self.formname.endswith('.xlsx'):
            pd.DataFrame(tup, columns=self.listheadingdata).to_excel(self.formname)
        #
        # res = open(self.formname, 'w')
        # wr = csv.writer(res, dialect='excel')
        # wr.writerow(self.listheadingdata)
        # for i in range(len(tup)):
        #     wr.writerow(tup[i])
        self.table1(tup, sql)
        if len(duplicate) == 0:
            self.showtooltip("Form generated sucessfully")
        else:
            self.showtooltip("Form generated sucessfully")
            QtGui.QMessageBox.warning(ui.enrolformFrame, "Message",
                                      "Form generated sucessfully\n Enrolment numbers : " + duplicate + " are not enrolled.",
                                      'OK')
        # res.close()
        ui.entryBox.setText("")
        os.startfile(self.formname)

    def picselect(self, obj):

        if obj.objectName() == 'selectpicturePushButton':
            self.candidphoto = QtGui.QFileDialog.getOpenFileName(ui.Enrol, "Select the candidate's picture", '.',
                                                                 filter="Images (*.png *.jpg *JPG *PNG)")
            if not self.candidphoto:
                self.candidphoto = ''
                return
            ui.selectpictureLabel.setPixmap(QtGui.QPixmap(self.candidphoto))

        if obj.objectName() == 'enrol_signaturePushButton':
            self.signaturephoto = QtGui.QFileDialog.getOpenFileName(ui.Enrol, "Select candiate's signature picture", '.',
                                                                    filter="Images (*.png *.jpg *.PNG *JPG)")
            if not self.signaturephoto:
                return
            ui.selectsignatureLabel.setPixmap(QtGui.QPixmap(self.signaturephoto))

    def check_enrol_form_data(self):

        proceed = True;

        if not ui.updateentryCheckBox.isChecked():
            sql = "select Student_Name from enrolment where Enrolment_Number='" + ui.enrolmentnumLineEdit.displayText().strip() + "'"

            try:
                tup = ENROLMENT_FORM.enroll().execute(sql)

            except Exception as e:
                if 'UNIQUE' in str(e) and 'Enrolment_Number' in str(e):
                    QtGui.QMessageBox.warning(ui.Enrol, 'Please use another enrolment number',
                                              '\nEnrolment number must be unique.\nSomeone already has the same enrolment number. If you want to update the present entry , then check the Update Entry check box.',
                                              'OK');

                elif 'UNIQUE' in str(e) and 'Aadhaar_Number' in str(e):
                    QtGui.QMessageBox.warning(ui.Enrol, 'Aadhaar number already exists',
                                              '\nAadhaar number must be unique.\nSomeone already has the same Aadhaar number. If you want to Update the Present Entry , then check the Update Entry check box.',
                                              'OK');
                return

        if not proceed:
            return

        def set_margin_red_style(obj):

            try:
                obj.setStyleSheet('border-color:red;border-width:2px;border-style:groove;')
                obj.setPlaceholderText('Mandatory field')
            except:
                pass;

        mailid = ui.emailLineEdit.displayText().strip()

        if len(mailid) and ('@' not in mailid or (mailid.rfind('.') - mailid.rfind('@')) < 2):
            QtGui.QMessageBox.warning(ui.Enrol, "Warning",
                                      "\n\nPlease make sure that the entered Email address is a valid one.", 'OK')

            return

        for i in [ui.enrolmentnumLineEdit, ui.fullnameLineEdit, ui.addressTextEdit, ui.unitLineEdit,
                  ui.aadhaarLineEdit]:

            if i == ui.addressTextEdit:
                if i.toPlainText() == '':
                    proceed = False;

                    set_margin_red_style(i)


            elif i.displayText().strip() == '':

                proceed = False
                set_margin_red_style(i)



            else:
                i.setStyleSheet('')
                i.setPlaceholderText('')

        if proceed:  # This runs only if all the non-null fields are filled.

            ui.addressTextEdit.setStyleSheet('')
            if len(ui.aadhaarLineEdit.displayText().strip()) != 12:
                proceed = False;
                QtGui.QMessageBox.warning(ui.Enrol, 'Warning',
                                          "\nValid Aadhaar Number should be 12 digits long.\n\n Please make sure that it's a valid 12 digit Aadhaar number",
                                          'OK')
                return;

            if len(ui.mobileLineEdit.displayText().strip()) and len(ui.mobileLineEdit.displayText().strip()) != 10:
                proceed = False;
                QtGui.QMessageBox.warning(ui.Enrol, 'Warning',
                                          '\n\n\nPlease Enter a Valid 10 digit Mobile number.',
                                          'OK');
                return


            else:
                proceed = True;
                self.get_enroll_form_data()


        else:
            QtGui.QMessageBox.warning(ui.Enrol, 'Warning',
                                      '\nMandatory fields should not be empty.\n\nMake sure that all the mandatory fields are filled.',
                                      'OK');

    def queryselectall(self):

        if ui.selectallCheckBox.isChecked():

            ui.selectallCheckBox.setStyleSheet("""
            color:chartreuse;

            font-size:16pt;

            font-family:'longdon decorative';

            font-weight:bold;

            text-decoration:underline;
            """)

            for i in ui.checkboxFrame.findChildren(QtGui.QCheckBox):

                if i.objectName() != 'selectallCheckBox':
                    i.setChecked(False);

                    i.setDisabled(True);

                    i.setStyleSheet("""

                    color:gray;

                    font-size:10pt;

                    font-weight:bold;

                    """)



        else:

            for i in ui.checkboxFrame.findChildren(QtGui.QCheckBox):
                i.setDisabled(False);

                i.setStyleSheet("#{}".format(i.objectName()) + """

                {
                color:white;

                font:14pt simonetta;

                font-weight:bold;

                }

                """ + "#{0}:hover".format(i.objectName()) + """

                {

                    text-decoration:underline;

                    font:15pt simonetta;

                    color:yellow;

                }

                """);

            ui.selectallCheckBox.setStyleSheet("""

#selectallCheckBox{

font: 75 16pt "longdon decorative";

color:rgb(255, 170, 0);

font-weight:bold;

}



#selectallCheckBox:hover{

font: 75 12pt "Georgia";

color:rgb(255, 148, 241);

font-weight:bold;

}

        """)

    def querycheckboxes(self, obj):
        if obj.isChecked():

            obj.setStyleSheet("""
            color:chartreuse;
            font-size:15.5pt;
            font-weight:bold;
            text-decoration:underline;
            font-family:'simonetta';""")
        else:

            obj.setStyleSheet("#{}".format(obj.objectName()) + """

{

color:white;
}

""" + "#{0}:hover".format(obj.objectName()) + """

{

    text-decoration:underline;
    color:yellow;

}""")

    def disable_query_checkbox_elements(self):

        for i in ui.enrolformFrame.findChildren(
                (QtGui.QLineEdit, QtGui.QComboBox, QtGui.QCheckBox, QtGui.QRadioButton, QtGui.QTextEdit)):
            i.setDisabled(True);

        for i in ui.bankformFrame.findChildren(QtGui.QLineEdit):
            i.setDisabled(True);

        for i in ui.instFrame.findChildren((QtGui.QLineEdit, QtGui.QComboBox)):
            i.setDisabled(True);
        ui.dateofbirthDateEdit.setDisabled(True)
        ui.enroldateDateEdit.setDisabled(True)
        ui.vegRadioButton.setDisabled(True)
        ui.nonvegRadioButton.setDisabled(True)
        ui.submitPushButton.hide()
        ui.updateentryCheckBox.hide()
        ui.enrol_campsListWidget.setDisabled(True)
        ui.selectpicturePushButton.setDisabled(True)
        ui.enrol_signaturePushButton.setDisabled(True)
        ui.juniorCheckBox.setDisabled(True)

    def enable_query_checkbox_elements(self):

        for i in ui.enrolformFrame.findChildren(
                (QtGui.QLineEdit, QtGui.QComboBox, QtGui.QCheckBox, QtGui.QRadioButton, QtGui.QTextEdit)):
            i.setDisabled(False);

        for i in ui.bankformFrame.findChildren(QtGui.QLineEdit):
            i.setDisabled(False);

        for i in ui.instFrame.findChildren((QtGui.QLineEdit, QtGui.QComboBox)):
            i.setDisabled(False);

        ui.dateofbirthDateEdit.setDisabled(False)
        ui.enroldateDateEdit.setDisabled(False)
        ui.vegRadioButton.setDisabled(False)
        ui.nonvegRadioButton.setDisabled(False)
        ui.submitPushButton.show()
        ui.updateentryCheckBox.show()
        ui.enrol_campsListWidget.setDisabled(False)
        ui.selectpicturePushButton.setDisabled(False)
        ui.enrol_signaturePushButton.setDisabled(False)
        ui.juniorCheckBox.setDisabled(False)

    def display(self, obj):  # this executes when the Search button is pressed
        self.clear_enrolment_form()
        if obj.objectName() == 'searchPushButton':
            if ui.searchbyfieldLineEdit.displayText().strip() == '':
                QtGui.QMessageBox.warning(ui.enrolformFrame, "Warning",
                                          "\n\nSearch field should not be empty while searching", 'OK')
                return;

            self.disable_query_checkbox_elements();

        obj = ENROLMENT_FORM.enroll()

        search_field_text = ui.searchbyfieldLineEdit.displayText().strip()

        if ui.aadhaarnumRadioButton.isChecked():
            self.field = 'Aadhaar_Number'
            # if len(search_field_text)==12:
            #     search_field_text = search_field_text[:4]+' '+search_field_text[4:8]+' '+search_field_text[8:]


        elif ui.enrolmentnumRadioButton.isChecked():
            self.field = 'Enrolment_Number'

        ui.enrolPushButton.setChecked(False)

        tuple = obj.search_by_field(self.field, search_field_text);

        if not tuple:
            self.showtooltip("No Details Found")
            ui.enrolPushButton.setChecked(True)
            self.enable_query_checkbox_elements()
            return

        candidatephoto = self.check_if_img_exists(tuple[0])
        if candidatephoto:
            ui.selectpictureLabel.setPixmap(QtGui.QPixmap(candidatephoto))

        signaturephoto = self.check_if_img_exists(tuple[0] + '_sign')
        if signaturephoto:
            ui.selectsignatureLabel.setPixmap(QtGui.QPixmap(signaturephoto))

        months = {'Jan': 1, 'Feb': 2, 'Mar': 3, 'Apr': 4, 'May': 5, 'Jun': 6, 'Jul': 7, 'Aug': 8,
                  'Sep': 9, 'Oct': 10, 'Nov': 11, 'Dec': 12}

        ui.enrolmentnumLineEdit.setText(tuple[0])
        ui.rankComboBox.setCurrentIndex(ui.rankComboBox.findText(tuple[1]))
        ui.aadhaarLineEdit.setText(str(tuple[2]))

        ui.fullnameLineEdit.setText(tuple[3])
        ui.SmiddlenameLineEdit.setText(tuple[4])
        ui.SlastnameLineEdit.setText(tuple[5])
        # FULL NAME IS tuple[6]



        ui.fathernameLineEdit.setText(tuple[7])
        ui.FmiddlenameLineEdit.setText(tuple[8])
        ui.FlastnameLineEdit.setText(tuple[9])
        # FUll father Name is tuple[10]



        ui.mothernameLineEdit.setText(tuple[11])
        ui.MmiddlenameLineEdit.setText(tuple[12])
        ui.MlastnameLineEdit.setText(tuple[13])
        # FULL MOTHER NAME is tuple[14]


        ui.sexComboBox.setCurrentIndex(ui.sexComboBox.findText(tuple[15]))

        dob = tuple[16].split('/')
        ui.dateofbirthDateEdit.setDate(QtCore.QDate(int(dob[0]), months[dob[1]], int(dob[2])))

        ui.addressTextEdit.setText(tuple[17])
        ui.emailLineEdit.setText(tuple[18])
        ui.mobileLineEdit.setText(str(tuple[19]))
        ui.bloodgroupComboBox.setCurrentIndex(ui.bloodgroupComboBox.findText(tuple[20]))
        ui.railwaystationLineEdit.setText(str(tuple[21]))
        ui.policestationLineEdit.setText(tuple[22])
        ui.educationLineEdit.setText(tuple[23])
        ui.marksLineEdit.setText(str(tuple[24]))
        ui.identificationmarksLineEdit.setText(tuple[25])
        ui.criminalcourtTextEdit.setText(tuple[26])
        ui.schoolorcollegeLineEdit.setText(tuple[27])
        ui.enrollpermissionLineEdit.setText(tuple[28])
        ui.earliercandidateComboBox.setCurrentIndex(ui.earliercandidateComboBox.findText(tuple[29]))
        ui.earlierenrolmentnumLineEdit.setText(tuple[30])
        ui.dismissedTextEdit.setText(tuple[31])

        if tuple[32] == "A":
            ui.AcertRadioButton.setChecked(True)
        elif tuple[32] == "B":
            ui.BcertRadioButton.setChecked(True)
        elif tuple[32] == "C":
            ui.CcertRadioButton.setChecked(True)
        else:
            ui.NullcertRadioButton.setChecked(True)

        camplist = tuple[33].split(',')

        for i in range(ui.enrol_campsListWidget.count()):
            if ui.enrol_campsListWidget.item(i).text() in camplist:
                ui.enrol_campsListWidget.item(i).setSelected(True)

        ui.extraactivitiesTextEdit.setText(tuple[34])
        ui.specialachievementsTextEdit.setText(tuple[35])

        enroldate = tuple[36].split('/')
        ui.enroldateDateEdit.setDate(QtCore.QDate(int(enroldate[0]), months[enroldate[1]], int(enroldate[2])))

        ui.remarksTextEdit.setText(tuple[37])
        if tuple[38] == "Veg":
            ui.vegRadioButton.setChecked(1)
        else:
            ui.nonvegRadioButton.setChecked(1)

        ui.banknameLineEdit.setText(tuple[39])
        ui.bankbranchLineEdit.setText(tuple[40])
        ui.accountnameLineEdit.setText(tuple[41])
        ui.accountnumLineEdit.setText(str(tuple[42]))
        ui.ifsccodeLineEdit.setText(tuple[43])
        ui.micrLineEdit.setText(str(tuple[44]))
        ui.pannumLineEdit.setText(str(tuple[45]))
        ui.institutionenrollComboBox.setCurrentIndex(ui.institutionenrollComboBox.findText(tuple[46]))
        ui.unitLineEdit.setText(tuple[47])
        if tuple[48]=="junior":
            ui.juniorCheckBox.setChecked(True)
        else:
            ui.juniorCheckBox.setChecked(False)
        self.showtooltip("Sucessfully Searched")

    def clear_enrolment_form(self):
        for i in ui.enrolformFrame.findChildren((QtGui.QLineEdit, QtGui.QTextEdit)):
            i.clear()
        for i in ui.bankformFrame.findChildren(QtGui.QLineEdit):
            i.clear()
        for i in ui.instFrame.findChildren((QtGui.QLineEdit)):
            i.clear()

        ui.enrol_campsListWidget.clearSelection()
        ui.aadhaarLineEdit.setPlaceholderText("Enter 12 digit aadhaar number")
        ui.fullnameLineEdit.setPlaceholderText("First Name")
        ui.unitLineEdit.setText('4KAR')
        self.candidphoto = ''
        self.signaturephoto = ''
        ui.selectpictureLabel.clear()
        ui.selectsignatureLabel.clear()

    def enrol_adhaar_radio_change(self):
        if ui.aadhaarnumRadioButton.isChecked():
            ui.searchbyfieldLineEdit.clear();
            ui.searchbyfieldLineEdit.setValidator(QtGui.QDoubleValidator())
            ui.searchbyfieldLineEdit.setMaxLength(12)

        elif ui.enrolmentnumRadioButton.isChecked():
            ui.searchbyfieldLineEdit.clear()
            ui.searchbyfieldLineEdit.setMaxLength(1000)
            ui.searchbyfieldLineEdit.setValidator(None)

    def get_enroll_form_data(self):
        enrolmentnum = ui.enrolmentnumLineEdit.displayText().strip();
        aadhaarnum = ui.aadhaarLineEdit.displayText().strip()
        rank = ui.rankComboBox.currentText()

        fullname = ui.fullnameLineEdit.displayText().strip()
        Smiddlename = ui.SmiddlenameLineEdit.displayText().strip()
        Slastname = ui.SlastnameLineEdit.displayText().strip()

        fathername = ui.fathernameLineEdit.displayText().strip()
        Fmiddlename = ui.FmiddlenameLineEdit.displayText().strip()
        Flastname = ui.FlastnameLineEdit.displayText().strip()

        mothername = ui.mothernameLineEdit.displayText().strip()
        Mmiddlename = ui.MmiddlenameLineEdit.displayText().strip()
        Mlastname = ui.MlastnameLineEdit.displayText().strip()

        sex = ui.sexComboBox.currentText()

        dateofbirth = ui.dateofbirthDateEdit.text().strip()

        address = ui.addressTextEdit.toPlainText()

        email = ui.emailLineEdit.displayText().strip()

        mobilenum = ui.mobileLineEdit.displayText().strip()

        bloodgroup = ui.bloodgroupComboBox.currentText()

        railwaystation=ui.railwaystationLineEdit.displayText().strip()

        policestation=ui.policestationLineEdit.displayText().strip()

        education=ui.educationLineEdit.displayText().strip()

        educationmarks=ui.marksLineEdit.displayText().strip()

        identificationmarks=ui.identificationmarksLineEdit.displayText().strip()

        criminalcourt=ui.criminalcourtTextEdit.toPlainText()

        nameofschoolcollege=ui.schoolorcollegeLineEdit.displayText().strip()

        enrolpermission=ui.enrollpermissionLineEdit.displayText().strip()

        earliercandidate=ui.earliercandidateComboBox.currentText()

        earlierenrolmentnum=ui.earlierenrolmentnumLineEdit.displayText().strip()

        dismissed=ui.dismissedTextEdit.toPlainText()


        bankname = ui.banknameLineEdit.displayText().strip()

        bankbranch = ui.bankbranchLineEdit.displayText().strip()

        accountname = ui.accountnameLineEdit.displayText().strip()

        accountnum = ui.accountnumLineEdit.displayText().strip()

        ifsccode = ui.ifsccodeLineEdit.displayText().strip()

        micr = ui.micrLineEdit.displayText().strip()

        pannum=ui.pannumLineEdit.displayText().strip()

        institutionname = ui.institutionenrollComboBox.currentText()

        unit = ui.unitLineEdit.displayText().strip()

        enrolldate = ui.enroldateDateEdit.text().strip()

        remarks = ui.remarksTextEdit.toPlainText()

        specialachievements = ui.specialachievementsTextEdit.toPlainText()

        extracurricularactivities = ui.extraactivitiesTextEdit.toPlainText()
        seniority="senior"
        if ui.juniorCheckBox.isChecked():
            seniority="junior"



        if ui.updateentryCheckBox.isChecked():
            img = self.check_if_img_exists(enrolmentnum);
            imgsign = self.check_if_img_exists(enrolmentnum+'_sign')
            if len(img) and len(self.candidphoto):
                try:
                    os.remove(img);
                except(FileNotFoundError):
                    if img.endswith('.png'):
                        os.remove(img.replace('png', 'PNG'))
                    elif img.endswith('.jpg'):
                        os.remove(img.replace('jpg', 'JPG'));
            if len(imgsign) and len(self.signaturephoto):
                try:
                    os.remove(imgsign) ;
                except(FileNotFoundError):
                    if imgsign.endswith('.png'):
                        os.remove(imgsign.replace('png','PNG'))
                    elif imgsign.endswith('.jpg'):
                        os.remove(imgsign.replace('jpg','JPG')) ;


        if self.candidphoto:
            ext = self.candidphoto[self.candidphoto.rfind('.') + 1:]
            copy2(self.candidphoto, "candidate photos\{}.{}".format(enrolmentnum, ext))

        if self.signaturephoto:
            ext = self.signaturephoto[self.signaturephoto.rfind('.') + 1:]
            copy2(self.signaturephoto, "candidate photos\{}_sign.{}".format(enrolmentnum, ext))

        campsattended = ui.enrol_campsListWidget.selectedItems()
        campsattended = '' if not campsattended else campsattended

        if campsattended:
            campsattended = ','.join([i.text() for i in campsattended])

        certificate = ""
        if ui.AcertRadioButton.isChecked():
            certificate = "A"
        if ui.BcertRadioButton.isChecked():
            certificate = "B"
        if ui.CcertRadioButton.isChecked():
            certificate = "C"
        Meal_Preference = "Veg"
        if ui.nonvegRadioButton.isChecked():
            Meal_Preference = "Nonveg"

        obj = ENROLMENT_FORM.enroll()

        if ui.updateentryCheckBox.isChecked():
            try:
                obj.delete_by_Enrolment('enrolment', enrolmentnum)
                obj.enrol_student(enrolmentnum, rank, aadhaarnum, fullname, Smiddlename, Slastname, fullname,fathername,

                  Fmiddlename, Flastname, fathername, mothername, Mmiddlename, Mlastname, mothername,sex, dateofbirth, address,

                  email, mobilenum, bloodgroup, railwaystation, policestation, education, educationmarks,
                  identificationmarks,criminalcourt,nameofschoolcollege,enrolpermission,earliercandidate,earlierenrolmentnum,
                  dismissed,certificate, campsattended, extracurricularactivities, specialachievements, enrolldate,
                  remarks, Meal_Preference, bankname, bankbranch,accountname,accountnum, ifsccode, micr,pannum, institutionname, unit,seniority)


            except (IntegrityError, OperationalError) as e:
                if 'UNIQUE' in str(e) and 'Aadhaar_Number' in str(e):
                    QtGui.QMessageBox.warning(ui.Enrol, 'Aadhaar number already exists',
                                              '\nAadhaar number must be unique.\nSomeone already has the same Aadhaar number. Please check the Aadhaar Number.',
                                              'OK');
                self.showtooltip("Update Failed !")
                return

            self.showtooltip("Updated successfully")


        else:
            try:
                obj.enrol_student(enrolmentnum, rank, aadhaarnum, fullname, Smiddlename, Slastname, fullname,fathername,
                      Fmiddlename, Flastname, fathername, mothername, Mmiddlename, Mlastname, mothername,sex, dateofbirth, address,
                      email, mobilenum, bloodgroup,railwaystation, policestation, education, educationmarks,
                      identificationmarks,criminalcourt,nameofschoolcollege,enrolpermission,earliercandidate,earlierenrolmentnum,
                      dismissed, certificate, campsattended, extracurricularactivities,specialachievements, enrolldate, remarks,
                      Meal_Preference, bankname, bankbranch,accountname,accountnum, ifsccode, micr,pannum, institutionname, unit,seniority)

            except (OperationalError, IntegrityError) as e:
                if 'UNIQUE' in str(e) and 'Enrolment_Number' in str(e):
                    QtGui.QMessageBox.warning(ui.Enrol, 'Please use another enrolment number',
                                              '\nEnrolment number must be unique.\nSomeone already has the same enrolment number. If you want to update the present entry , then check the Update Entry check box.',
                                              'OK')

                elif 'UNIQUE' in str(e) and 'Aadhaar_Number' in str(e):
                    QtGui.QMessageBox.warning(ui.Enrol, 'Aadhaar number already exists',
                                              '\nAadhaar number must be unique.\nSomeone already has the same Aadhaar number. If you want to Update the Present Entry , then check the Update Entry check box.',
                                              'OK')
                return

            self.showtooltip("Inserted successfully")

        self.clear_enrolment_form()

        con = connect("ncc.db")
        data = pd.read_sql("select * from enrolment", con)
        if self.check_if_img_exists(enrolmentnum) == "":
            im = ""
        else:
            im = pdfImage(self.check_if_img_exists(enrolmentnum))
            im.drawWidth = 100
            im.drawHeight = 100
        lis = [
            ['','','',''],
            ['Enrolment Number', ':', enrolmentnum, '', im],
            ['Rank', ':', rank],
            ['Aadhaar Number', ':', aadhaarnum],
            ['Student Name', ':', fullname, Smiddlename, Slastname],
            ['Fathers Name', ':', fathername, Fmiddlename, Flastname],
            ['Mothers Name', ':', mothername, Mmiddlename, Mlastname],
            ['Sex', ':', sex],
            ['Date Of Birth', ':', dateofbirth],
            ['Address', ':', address],
            ['Email', ':', email],
            ['Mobile', ':', mobilenum],
            ['Blood Group', ':', bloodgroup],
            ['Nearest Railway Station', ':', railwaystation],
            ['Nearest Police Station', ':', policestation]

            ]

        lis1=[
            ['Education Qualifications\nand Marks in %', ':',education, educationmarks],
            ['Identification Marks', ':', identificationmarks],
            [
                'Have you ever been\nconvinced by a criminal\ncourt and if so\nin what circumstances and\nwhat was the sentence?\nAttach relevant documents',
                ':', criminalcourt],
            ['Name of the School\nor College and Stream', ':', nameofschoolcollege],
            ['Willing to be enrolled\nand undergo training\nunder the National\nCadet Corps Act. 1948.', ':',
             enrolpermission],
            ['Have you been enrolled in\nNCC earlier.If Yes your\nEnrolment Number', ':', earliercandidate,
             earlierenrolmentnum],
            [
                'Have you been dismissed\nfrom NCC/The\nTerritorial Army/\nThe Indian Armed Forces.\nplease provide details',
                ':', dismissed],
            ['Certificate', ':', certificate],
            ['Camps Attended', ':', campsattended],
            ['Extra Activities', ':', extracurricularactivities],
            ['Achievements', ':', specialachievements],
            ['Enroll Date', ':', enrolldate]

        ]



        lis2=[
            ['Remarks', ':', remarks],
            ['Meal Preferences', ':', Meal_Preference],
            ['Bank Name', ':', bankname],
            ['Branch', ':', bankbranch],
            ['Account Name', ':', accountname],
            ['Account Number', ':', accountnum],
            ['IFSC Code', ':', ifsccode],
            ['MICR', ':', micr],
            ['PAN Number\n(if allotted)', ':', pannum],
            ['Institution', ':', institutionname],
            ['UNIT', ':', unit],
            ['Seniority', ':', seniority]
        ]

        from reportlab.pdfbase.ttfonts import TTFont
        from reportlab.pdfbase import pdfmetrics
        doc = SimpleDocTemplate("candidate photos\{}_pdf.{}".format(enrolmentnum, "pdf"), pagesize=A4,
                                rightMargin=20, leftMargin=20,
                                topMargin=45, bottomMargin=20)
        Story = []
        pdfmetrics.registerFont(TTFont('longdon',os.path._getfullpathname('Longdon_Decorative.ttf')))
        pdfmetrics.registerFont(TTFont('caladea', os.path._getfullpathname('CALADEA-REGULAR.ttf')))
        pdfmetrics.registerFont(TTFont('caladea-bold', os.path._getfullpathname('CALADEA-BOLD.ttf')))

        styles = getSampleStyleSheet()
        styles.add(ParagraphStyle(name='rightallignment', alignment=TA_RIGHT, fontSize=13, fontName='caladea'))
        styles.add(ParagraphStyle(name='leftallignment', alignment=TA_LEFT, fontSize=13, fontName='caladea'))
        styles.add(ParagraphStyle(name='heading', alignment=TA_CENTER, fontName='caladea-bold'))
        styles.add(ParagraphStyle(name='normal', fontSize=11.5, leading=15, fontName='caladea'))
        styles.add(ParagraphStyle(name='list', fontSize=11.5, fontName='caladea'))
        styles.add(ParagraphStyle(name='mainheading', fontSize=11.5, fontName='longdon' , alignment=TA_CENTER))
        styles.add(ParagraphStyle(name='header', fontSize=9, alignment=TA_RIGHT))

        t = Table(lis, colWidths=[2 * inch, 0.1 * inch, 2 * inch, 2 * inch, 2 * inch], hAlign='LEFT')
        t.setStyle(TableStyle([('TOPPADDING', (0, 0), (-1, -1), 20),
                               ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                               ('VALIGN', (0, 0), (-1, 0), 'MIDDLE'),
                               ('FONT', (0, 0), (-1, -1), 'caladea')
                               ]))


        t1 = Table(lis1, colWidths=[2 * inch, 0.1 * inch, 2 * inch, 2 * inch, 2 * inch], hAlign='LEFT')
        t1.setStyle(TableStyle([('TOPPADDING', (0, 0), (-1, -1), 20),
                               ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                               ('VALIGN', (0, 0), (-1, 0), 'MIDDLE'),
                                ('FONT', (0, 0), (-1, -1), 'caladea')
                                ]))
        t2 = Table(lis2, colWidths=[2 * inch, 0.1 * inch, 2 * inch, 2 * inch, 2 * inch], hAlign='LEFT')
        t2.setStyle(TableStyle([('TOPPADDING', (0, 0), (-1, -1), 20),
                               ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                               ('VALIGN', (0, 0), (-1, 0), 'MIDDLE'),
                                ('FONT', (0, 0), (-1, -1), 'caladea')
                                ]))

        Story.append(Paragraph(enrolmentnum + " ~ Page - 1", styles["header"]))
        Story.append(Paragraph(
            "____________________________________________________________________________________________________________",
            styles["header"]))

        Story.append(Paragraph("<font size='25'>ENROLMENT FORM</font>", styles["mainheading"]))
        Story.append(t)
        Story.append(PageBreak())
        Story.append(Paragraph(enrolmentnum+" ~ Page - 2",styles["header"]))
        Story.append(Paragraph("____________________________________________________________________________________________________________", styles["header"]))
        Story.append(t1)
        Story.append(PageBreak())
        Story.append(Paragraph(enrolmentnum + " ~ Page - 3", styles["header"]))
        Story.append(Paragraph("____________________________________________________________________________________________________________",styles["header"]))

        Story.append(t2)
        Story.append(Spacer(1, 12))

        Story.append(
            Paragraph("@<br/><br/>Place:_______________<br/><br/>Date:______________", styles["leftallignment"]))
        Story.append(Paragraph("Signature of the Applicant", styles["rightallignment"]))
        Story.append(PageBreak())
        Story.append(Paragraph(enrolmentnum + " ~ Page - 4", styles["header"]))
        Story.append(Paragraph(
            "____________________________________________________________________________________________________________",
            styles["header"]))


        Story.append(
            Paragraph("<font size='15'><u>DECLARATION ON ACCEPTANCE FOR ENROLMENT</u></font>", styles["heading"]))
        Story.append(Spacer(1, 40))
        Story.append(Paragraph(
            "1. I solemnly declare that the answers I have given to the questions in this form are true and no part of them is false, and that I am willing to fulfill the engagement made.",
            styles["normal"]))
        Story.append(Spacer(1, 10))
        Story.append(Paragraph(
            "2. I <font size=13><b><u>"+fullname.upper()+"</u></b></font> promise that I will honestly and faithfully serve my country and abide by the Rules and Regulations of the National Cadet Corps that I will to the best of my ability, attended all parades and camps as may be required by the Commanding Officer from time to time..",
            styles["normal"]))
        Story.append(Spacer(1, 10))
        Story.append(Paragraph(
            "3. I <font size=13><b><u>"+fullname.upper()+"</u></b></font> further promise that after enrolment, I will have no claim on authorities for any compensation and the event of injury or death due to accident during training camps, courses, traveling and while on YEP or any other such NCC events like RDC and IDC.",
            styles["normal"]))
        Story.append(Spacer(1, 10))
        Story.append(Paragraph(
            "4. I hereby certify that I have not been enrolled myself in NCC as a SD / SW, JD / JW cadets in Army /Navy / Air force.Earlier I have not appeared for A, B and C certificate examination.",
            styles["normal"]))
        Story.append(
            Paragraph("@<br/><br/>Place:_______________<br/><br/>Date:______________", styles["leftallignment"]))
        Story.append(Paragraph("Signature of the Applicant", styles["rightallignment"]))
        Story.append(Spacer(1, 80))


        Story.append(Paragraph("<font size='15'><u>DECLARATION BY PARENT/GUARDIAN</u></font>", styles["heading"]))
        Story.append(Spacer(1, 40))
        Story.append(Paragraph(
            "1. I solemnly declare that the answer given in this form are true and that no part of them is false, and that My Son / Daughter / Ward is willing to fulfill the engagement made.",
            styles["normal"]))
        Story.append(Spacer(1, 10))
        Story.append(Paragraph(
            "2. I ___________________________________ promise that after enrolment of my Son / Daughter / Ward. I will have no claim on authorities for any compensation in the event of any injury or death due to accident during training camps, courses, traveling and while on YEP or any other such NCC events like RDC and IDC",
            styles["normal"]))
        Story.append(Spacer(1, 10))
        Story.append(
            Paragraph("@<br/><br/>Place:_______________<br/><br/>Date:______________", styles["leftallignment"]))
        Story.append(Paragraph("Signature of the Parent / GUARDIAN", styles["rightallignment"]))
        Story.append(PageBreak())
        Story.append(Paragraph(enrolmentnum + " ~ Page - 5", styles["header"]))
        Story.append(Paragraph(
            "____________________________________________________________________________________________________________",
            styles["header"]))

        Story.append(Paragraph("<font size='15'><u>CERTIFICATE</u></font>", styles["heading"]))
        Story.append(Spacer(1, 20))
        Story.append(
            Paragraph("1. Certified that the applicant understands and agrees to the conditions of enrolment.",
                      styles["normal"]))
        Story.append(Spacer(1, 10))
        Story.append(Paragraph(
            "2. Certified that the applicant and his Parent / Guardian understand and agree to the condition on enrolment.",
            styles["normal"]))
        Story.append(Spacer(1, 10))
        Story.append(Paragraph("@<br/><br/>Place:_______________<br/><br/>Date:______________<br/><br/>(Unit seal)",
                               styles["leftallignment"]))
        Story.append(Spacer(1, 10))
        Story.append(Paragraph("* For Minors only. Score out inapplicable portion.", styles["normal"]))
        Story.append(Paragraph("Signature of the Parent / Guardian", styles["rightallignment"]))
        Story.append(Spacer(1, 20))
        Story.append(
            Paragraph("<font size='15'><u>CERTIFICATE BY PRINCIPAL / HEAD MASTER</u></font>", styles["heading"]))
        Story.append(Spacer(1, 10))
        Story.append(Paragraph("I am satisfied that the applicant is in order and the application fulfills the condition of enrolment and that he / she is suitable for enrolment in the unit, may be medically examined.",styles["normal"]))
        Story.append(Spacer(1, 10))
        Story.append(
            Paragraph("@<br/><br/>Place:_______________<br/><br/>Date:______________", styles["leftallignment"]))
        Story.append(Paragraph("Signature of the Principal / Head Master", styles["rightallignment"]))
        Story.append(Spacer(1, 20))
        Story.append(
            Paragraph("<font size='15'><u>TO BE COMPLETED BY MEDICAL OFFICER<br/><br/>BEFORE ENROLMENT</u></font>",
                      styles["heading"]))
        Story.append(Spacer(1, 30))
        Story.append(Paragraph(
            "1. I have examined(Name) _______________________________________________________ on ___________________________________ (date) and consider him / her, fit/unfit for enrolment as a Cadet in the National Cadet Corps",
            styles["normal"]))
        Story.append(Spacer(1, 10))
        Story.append(
            Paragraph("2. He / She has been inoculated and vaccinated against the following:", styles["normal"]))
        Story.append(Spacer(1, 10))
        Story.append(Paragraph(" a. Typhoid(TAB)", styles['list']))
        Story.append(Paragraph(" b. Tetanus", styles['list']))
        Story.append(Paragraph(" c. Tuberculosis(BCG)", styles['list']))
        Story.append(Spacer(1, 10))
        Story.append(
            Paragraph("@<br/><br/>Place:_______________<br/><br/>Date:______________", styles["leftallignment"]))
        Story.append(Paragraph(
            "Signature:____________________<br/><br/>Designation:____________________<br/><br/>(Medical Officer)",
            styles["rightallignment"]))
        Story.append(PageBreak())
        Story.append(Paragraph(enrolmentnum + " ~ Page - 6", styles["header"]))
        Story.append(Paragraph(
            "____________________________________________________________________________________________________________",
            styles["header"]))

        Story.append(
            Paragraph("<font size='15'><u>TO BE USED FOR EXTENSION OF ENROLMENT<br/><br/>(See Rules 13)</u></font>",
                      styles["heading"]))
        Story.append(Spacer(1, 40))
        if not ui.juniorCheckBox.isChecked():
            Story.append(Paragraph(
                "A. I agree to extend my enrolment for one year and am willing to fulfill the engagement made.",
                styles["normal"]))
            Story.append(Spacer(1, 10))
            Story.append(Paragraph("@<br/><br/>Place:_______________<br/><br/>Date:______________<br/><br/>Confirmed",
                                   styles["leftallignment"]))
            Story.append(Paragraph("Signature of the Parent / Guardian", styles["rightallignment"]))
            Story.append(Spacer(1, 30))
            Story.append(
                Paragraph("@<br/><br/>Place:_______________<br/><br/>Date:______________", styles["leftallignment"]))
            Story.append(Paragraph("Signature of the Commanding Officer", styles["rightallignment"]))
            Story.append(Spacer(1, 60))
            Story.append(Paragraph(
                "B. I agree to extend enrolment of my Son / Daughter / Ward for one year and am willing to fulfill the engagement made.",
                styles["normal"]))
        else:
            Story.append(Paragraph(
                "My Son / Daughter / Ward agrees to extend the enrolment for one year and is willing to fulfill the engagement made.",
                styles["normal"]))
        Story.append(Spacer(1, 30))
        Story.append(Paragraph("@<br/><br/>Place:_______________<br/><br/>Date:______________<br/><br/>Confirmed",
                               styles["leftallignment"]))
        Story.append(Paragraph("Signature of the Parent /Guardian", styles["rightallignment"]))
        Story.append(Spacer(1, 30))
        Story.append(Paragraph(
            "@<br/><br/>Place:_______________<br/><br/>Date from which extension start:______________<br/><br/>(Unit Seal)",
            styles["leftallignment"]))
        Story.append(Paragraph("Signature of the Head Master", styles["rightallignment"]))
        Story.append(Spacer(1, 30))
        Story.append(Paragraph("NOTE: * This form will be retained in the school in which the unit is located",
                               styles["normal"]))

        """Story.append(PageBreak())"""
        doc.build(Story)

        try:

            if Smiddlename == "":
                Smiddlename = "-"
            data.to_csv(r'All candidate details.csv', float_format="%s", index=False)

        except(PermissionError):
            print("The csv file is already open. It needs to be closed before updating it.")

    def table1(self, res, msg):

        html3 = """

        <html>

        <head>

        <style type="text/css">



        table {

        border-collapse: collapse;

        text-align-last: center;

        vertical-align: middle;

        }

        table, td ,th{

            border: 1px solid black;

            text-align-last: center;

            height: 40px;

            text-align:center;



        }

        td{
            font-family:simonetta

            margin:3px 6px;

            text-align:center;

        }

        tr{

            background-color: white;

            text-align-last: center;

        }



        th{

            background-color: #7100b2;

            text-align-last: center;

            font-size: 120%;

            color: white;

        }

        """

        html3 = html3 + """</style>\n</head>\n<body>\n<table>\n<thead>\n"""

        msg1 = msg[7: msg.find('from')]


        i = 0

        mmsg = msg1.split(',')
        for i in mmsg:
            html3 = html3 + "\n<th>" + str(i) + "</th>\n"

        l=len(self.settings.value(ui.formsComboBox.currentText().replace(' ','_')+"_sql_fieldslist").split(',,,'))
        html3 = html3 + "\n<thead>\n"
        for i in range(len(res)):

            html3 = html3 + "<tr>\n"

            for j in range(l):
                html3 = html3 + "<td>" + str(res[i][j]) + "</td>\n"

            html3 = html3 + "</tr>\n"

        html3 = html3 + "</table>\n</body>\n</html>"

        ui.webView_2.setHtml(html3)


    def table(self, res, msg):
        html3 = """

        <html>

        <head>

        <style type="text/css">



        table {

        border-collapse: collapse;

        text-align-last: center;

        vertical-align: middle;

        }

        table, td ,th{

            border: 1px solid black;

            text-align-last: center;

            height: 30px;

            text-align:center;

        }

        td{
            margin:3px 6px;
            font-family:cambria;
            font-size:13pt;
            background-color:white;
            text-align:center;
            opacity:0.8;
            font-weight:bold;

        }

        tr{

            background-color: white;
            text-align-last: center;

        }



        th{

            background-color: rgb(177, 51, 255);
            opacity:0.8;
            font-family:'gabriola';
            text-align-last: center;
            font-size: 22px;
            height:20px;
            color: white;

        }

        """

        html3 = html3 + """</style>\n</head>\n<body>\n<table>\n<thead>\n<tr>\n"""

        msg1 = msg[7: msg.find('from')]


        i = 0
        mmsg = msg1.split(',')

        self.queryheading = mmsg

        for i in mmsg:
            html3 = html3 + "\n<th>" + str(i) + "</th>\n"


        html3 = html3 + "</tr>\n</thead>\n"

        for i in range(len(res)):

            html3 = html3 + "<tr>\n"

            for j in range(len(res[i])):
                html3 = html3 + "<td>" + str(res[i][j]) + "</td>\n"

            html3 = html3 + "</tr>\n"

        html3 = html3 + "</table>\n</body>\n</html>"

        ui.webView.setHtml(html3)

    def casemaker(self, sql2, field):
        if field == "Address" or field == "Education" or field == "Identification_mark" or field == "Criminal_Court" or field == "Name_of_School_College" or \
                        field == "Earlier_candidate" or field == "Earlier_Enrolment_Number" or field == "Camps_Attended" or field == "Dismissed" or\
                        field == "Nearest_Railway_Station" or field == "Nearest_Police_Station":
            length = 0
            l = 1
            sql33 = sql2
            while l > -1:
                l = sql33.find(field)
                if l > -1:
                    length = length + 1
                sql33 = sql33[l + 1:]
            while length > 0:
                str1 = sql2[sql2.find(field + "=") + len(field) + 2:]
                strr = str1[:str1.find("\"")]
                s = sql2[:sql2.find(field + "=") + len(field)]
                q = str1[str1.find("\"") + 1:]
                sql2 = s + " LIKE '%" + strr + "%' collate utf8_general_ci" + q
                length = length - 1
        else:
            length = 0
            l = 1
            sql33 = sql2
            while l > -1:
                l = sql33.find(field)
                if l > -1:
                    length = length + 1
                sql33 = sql33[l + 1:]
            while length > 0:
                str1 = sql2[sql2.find(field + "=") + len(field) + 2:]
                strr = str1[:str1.find("\"")]
                s = sql2[:sql2.find(field + "=") + len(field)]
                q = str1[str1.find("\"") + 1:]
                sql2 = s + " LIKE '" + strr + "' collate utf8_general_ci" + q
                length = length - 1
        return (sql2)

    def conquery(self):
        self.querytupple = []
        sql = """"""

        if ui.selectallCheckBox.isChecked():

            sql = "*"

        else:

            if ui.enrolmentCheckBox.isChecked(): sql += 'Enrolment_Number,';

            if ui.rankCheckBox.isChecked(): sql += 'Rank,'

            if ui.aadhaarCheckBox.isChecked(): sql += 'Aadhaar_Number,'

            if ui.sfnameCheckBox.isChecked(): sql += 'Student_Name,'

            if ui.ffnameCheckBox.isChecked(): sql += 'Fathers_Name,'

            if ui.mfnameCheckBox.isChecked(): sql += 'Mothers_Name,'

            if ui.dateofbirthCheckBox.isChecked(): sql += 'Date_Of_Birth,'

            if ui.sexCheckBox.isChecked(): sql += 'Sex,'

            if ui.bloodgroupCheckBox.isChecked(): sql += 'Blood_Group,'

            if ui.policestationCheckBox.isChecked():sql+='Nearest_Police_Station,'

            if ui.educationCheckBox.isChecked():sql+="Education,"

            if ui.schoolcollegeCheckBox.isChecked():sql+="Name_of_School_College,"

            if ui.pannumCheckBox.isChecked():sql+="Pan_Number,"

            if ui.earliercandidateCheckBox.isChecked():sql+="Earlier_candidate,"

            if ui.seniorityCheckBox.isChecked():sql+="Seniority,"

            if ui.extraCurricularActivitiesCheckBox.isChecked(): sql += 'Extra_Curricular_Activities,'

            if ui.specialAchievementsCheckBox.isChecked(): sql += 'Special_Achievements,'

            if ui.remarksCheckBox.isChecked(): sql += 'Remarks,'

            if ui.enrollDateCheckBox.isChecked(): sql += 'Enrol_Date,'

            if ui.vegitarianCheckBox.isChecked(): sql += 'Meal_Preference,'

            if ui.addressCheckBox.isChecked(): sql += 'Address,'

            if ui.emailCheckBox.isChecked(): sql += 'Email,'

            if ui.mobileCheckBox.isChecked(): sql += 'Mobile_Number,'

            if ui.banknameCheckBox.isChecked(): sql += 'Bank_Name,'

            if ui.bankbranchCheckBox.isChecked(): sql += 'Branch,'

            if ui.accountnameCheckBox.isChecked(): sql += 'Account_Name,'

            if ui.accountnumCheckBox.isChecked(): sql += 'Account_Number,'

            if ui.ifsccodeCheckBox.isChecked(): sql += 'IFSC_Code,'

            if ui.institutionCheckBox.isChecked(): sql += 'Institution,'

            if ui.unitCheckBox.isChecked(): sql += 'Unit,'

            if ui.micrCheckBox.isChecked(): sql += 'MICR,'

            if ui.campsAttendedCheckBox.isChecked(): sql += 'Camps_Attended'

            sql = sql.strip()

            if len(sql) and sql[-1] == ',': sql = sql[0:-1];

            if not sql:
                sql = '*'

                ui.selectallCheckBox.setChecked(True);

        sql1 = str(ui.conditionsentrylabel.text().strip())

        if sql == '*':

            if sql1 != "":

                sql = """select """+self.settings.value('enrolmentfields').replace(',,,',',')+""" from enrolment where """
                sql2 = ui.conditionsentrylabel.text()
                sql2 = self.casemaker(sql2, "Camps_Attended")
                sql2 = self.casemaker(sql2, "Enrolment_Number")
                sql2 = self.casemaker(sql2, "Address")
                sql2 = self.casemaker(sql2, "Student_Name")
                sql2 = self.casemaker(sql2, "Fathers_Name")
                sql2 = self.casemaker(sql2, "Mothers_Name")
                sql2 = self.casemaker(sql2, "Email")
                sql2 = self.casemaker(sql2, "Bank_Name")
                sql2 = self.casemaker(sql2, "Branch")
                sql2 = self.casemaker(sql2, "Account_Name")
                sql2 = self.casemaker(sql2, "Unit")
                sql2 = self.casemaker(sql2, "Education")
                sql2 = self.casemaker(sql2, "Identification_mark")
                sql2 = self.casemaker(sql2, "Criminal_Court")
                sql2 = self.casemaker(sql2, "Name_of_School_College")
                sql2 = self.casemaker(sql2, "Earlier_candidate")
                sql2 = self.casemaker(sql2, "Earlier_Enrolment_Number")
                sql2 = self.casemaker(sql2, "Dismissed")
                sql2 = self.casemaker(sql2, "Nearest_Railway_Station")
                sql2 = self.casemaker(sql2, "Nearest_Police_Station")




                sql = sql + sql2

            else:

                sql = """select """+self.settings.value('enrolmentfields').replace(',,,',',')+""" from enrolment"""


        else:
            sql2 = ui.conditionsentrylabel.text()
            sql2 = self.casemaker(sql2, "Camps_Attended")
            sql2 = self.casemaker(sql2, "Enrolment_Number")
            sql2 = self.casemaker(sql2, "Address")
            sql2 = self.casemaker(sql2, "Student_Name")
            sql2 = self.casemaker(sql2, "Fathers_Name")
            sql2 = self.casemaker(sql2, "Mothers_Name")
            sql2 = self.casemaker(sql2, "Email")
            sql2 = self.casemaker(sql2, "Bank_Name")
            sql2 = self.casemaker(sql2, "Branch")
            sql2 = self.casemaker(sql2, "Account_Name")
            sql2 = self.casemaker(sql2, "Unit")
            sql2 = self.casemaker(sql2, "Education")
            sql2 = self.casemaker(sql2, "Identification_mark")
            sql2 = self.casemaker(sql2, "Criminal_Court")
            sql2 = self.casemaker(sql2, "Name_of_School_College")
            sql2 = self.casemaker(sql2, "Earlier_candidate")
            sql2 = self.casemaker(sql2, "Earlier_Enrolment_Number")
            sql2 = self.casemaker(sql2, "Dismissed")
            sql2 = self.casemaker(sql2, "Nearest_Railway_Station")
            sql2 = self.casemaker(sql2, "Nearest_Police_Station")

            sql = ("select " + sql + " from enrolment where " + str(
                sql2)) if sql1 != "" else "select " + sql + " from enrolment "

        if sql[7] == "*":
            sql = """select """+self.settings.value('enrolmentfields').replace(',,,',',') + sql[9:len(sql)]
        self.querytupple = ENROLMENT_FORM.enroll().execute(sql)
        if len(self.querytupple) < 1:
            self.showtooltip("No Data Found")
            ui.webView.setHtml("")
            return

        self.query_hide_elements()
        self.table(self.querytupple, sql)
        self.showtooltip("Query is Sucessfull")

    def conback(self):

        sql = ui.conditionsentrylabel.text().strip().strip()
        if not len(sql):
            return
        ch = sql[-1]

        if sql[-2:] == 'or':
            ui.conditionsentrylabel.setText(sql[0:-2].strip())

        elif sql[-3:] == 'and':
            ui.conditionsentrylabel.setText(sql[0:-3].strip())



        elif ch in '(=)':

            ui.conditionsentrylabel.setText(sql[0:len(sql) - 1])

        elif ch == "\"":

            ui.conditionsentrylabel.setText(sql[0:str(sql[0:-1]).rfind('"') - 1])

            self.conback();

        else:

            sql = sql.strip()

            if sql.rfind(' ') > 0:
                ui.conditionsentrylabel.setText(sql[0:sql.rfind(' ')])

            else:
                ui.conditionsentrylabel.setText('');

    def coninsert(self):

        ch = ui.conditionlistcombobox.currentText()

        if ch == "Rank":
            ch1 = ui.rankqueryComboBox.currentText()
        elif ch == "Institution":
            ch1 = ui.institutionqueryComboBox.currentText()
        elif ch == "Blood_Group":
            ch1 = ui.bloodgroupqueryComboBox.currentText()
        elif ch == "Sex":
            ch1 = ui.sexqueryComboBox.currentText()
        elif ch == "Enrol_Date" or ch == "Date_Of_Birth":
            ch1 = ui.datequeryDateEdit.text()
        elif ch == "Certificate":
            ch1 = ui.certificatequeryComboBox.currentText()
        elif ch == "Meal_Preference":
            ch1 = ui.vegitarianqueryComboBox.currentText().replace(' ','_')
        elif ch == "Camps_Attended":
            ch1 = ui.campsattendedqueryComboBox.currentText()
        elif ch == "Seniority":
            ch1 = ui.seniorityqueryComboBox.currentText()
        else:
            ch1 = ui.valuelineEdit.text().strip()

        ch2 = ui.conditionsentrylabel.text().strip().strip() + ' '

        if (len(ch2) > 3):
            if (ch2[len(ch2) - 1] == 'r' and ch2[len(ch2) - 3] == ')') or (
                            ch2[len(ch2) - 1] == 'd' and ch2[len(ch2) - 4] == ')'):
                ch2 = ch2 + ' '


            if(ch2[len(ch2)-1]==" " and (ch2[len(ch2)-2]=="(" or ch2[len(ch2)-2]==")" or (ch2[len(ch2)-2]=="r" and ch2[len(ch2)-3]=="o") or (ch2[len(ch2)-2]=="d" and ch2[len(ch2)-3]=="n"))):
                print()





            else:
                self.showtooltip("Please use \'AND\' or \'OR\' between two conditions")

                return

        if ch != "Select Fields":

            if (ch1 != ""):

                if ch == "Aadhaar_Number" or ch == "Mobile_Number" or ch == "Account_number":

                    if ch1.isdigit():

                        if ch == "Mobile_Number":

                            if len(ch1) == 10:

                                ui.conditionsentrylabel.setText(
                                    ui.conditionsentrylabel.text().strip() + ' ' + ch + "=\"" + ch1 + "\"")

                            else:
                                self.showtooltip("Mobile number should contains 10 numbers")

                        elif ch == "Aadhaar_Number":

                            if len(ch1) == 12:

                                ui.conditionsentrylabel.setText(
                                    ui.conditionsentrylabel.text().strip() + ' ' + ch + "=\"" + ch1 + "\"")

                            else:
                                self.showtooltip("Aadhaar number should contains 12 numbers")


                        elif ch == "Account_Number":
                            if len(ch1) == 16:

                                ui.conditionsentrylabel.setText(
                                    ui.conditionsentrylabel.text().strip() + ' ' + ch + "=\"" + ch1 + "\"")

                            else:
                                self.showtooltip("Account number should contains 16 numbers")
                        else:

                            ui.conditionsentrylabel.setText(
                                ui.conditionsentrylabel.text().strip() + ' ' + ch + "=\"" + ch1 + "\"")

                    else:
                        self.showtooltip(ch + ' must contains only integral values.')

                else:

                    ui.conditionsentrylabel.setText(
                        ui.conditionsentrylabel.text().strip().strip() + ' ' + ch + "=\"" + ch1 + "\"")

            else:
                self.showtooltip("Please enter the values for the field selected.")

        else:
            self.showtooltip("Please select any one of the fields.")

        ui.valuelineEdit.clear()

def set_up_font():
    global fontDB
    fontDB = QtGui.QFontDatabase()
    fontDB.addApplicationFont(':/fonts/fonts/ALGER.TTF')
    fontDB.addApplicationFont(':/fonts/fonts/Ambient_Medium.ttf')
    fontDB.addApplicationFont(':/fonts/fonts/ARIAL.TTF')
    fontDB.addApplicationFont(':/fonts/fonts/ARIALBD.TTF')
    fontDB.addApplicationFont(':/fonts/fonts/ARIALBI.TTF')
    fontDB.addApplicationFont(':/fonts/fonts/ARIALI.TTF')
    fontDB.addApplicationFont(':/fonts/fonts/ARIALN.TTF')
    fontDB.addApplicationFont(':/fonts/fonts/ARIALNB.TTF')
    fontDB.addApplicationFont(':/fonts/fonts/ARIALNBI.TTF')
    fontDB.addApplicationFont(':/fonts/fonts/ARIALNI.TTF')
    fontDB.addApplicationFont(':/fonts/fonts/ARIBLK.TTF')
    fontDB.addApplicationFont(':/fonts/fonts/BASKVILL.TTF')
    fontDB.addApplicationFont(':/fonts/fonts/CALADEA-BOLD.TTF')
    fontDB.addApplicationFont(':/fonts/fonts/CALADEA-BOLDITALIC.TTF')
    fontDB.addApplicationFont(':/fonts/fonts/CALADEA-ITALIC.TTF')
    fontDB.addApplicationFont(':/fonts/fonts/CALADEA-REGULAR.TTF')
    fontDB.addApplicationFont(':/fonts/fonts/CALIBRI.TTF')
    fontDB.addApplicationFont(':/fonts/fonts/CALIBRIB.TTF')
    fontDB.addApplicationFont(':/fonts/fonts/CALIBRII.TTF')
    fontDB.addApplicationFont(':/fonts/fonts/CALIBRIL.TTF')
    fontDB.addApplicationFont(':/fonts/fonts/CALIBRILI.TTF')
    fontDB.addApplicationFont(':/fonts/fonts/CALIBRIZ.TTF')
    fontDB.addApplicationFont(':/fonts/fonts/CALIFB.TTF')
    fontDB.addApplicationFont(':/fonts/fonts/CALIFI.TTF')
    fontDB.addApplicationFont(':/fonts/fonts/CALIFR.TTF')
    fontDB.addApplicationFont(':/fonts/fonts/CAMBRIA.TTC')
    fontDB.addApplicationFont(':/fonts/fonts/CAMBRIAB.TTF')
    fontDB.addApplicationFont(':/fonts/fonts/CAMBRIAI.TTF')
    fontDB.addApplicationFont(':/fonts/fonts/CAMBRIAZ.TTF')
    fontDB.addApplicationFont(':/fonts/fonts/CANDARA.TTF')
    fontDB.addApplicationFont(':/fonts/fonts/CANDARAB.TTF')
    fontDB.addApplicationFont(':/fonts/fonts/CANDARAI.TTF')
    fontDB.addApplicationFont(':/fonts/fonts/CANDARAZ.TTF')
    fontDB.addApplicationFont(':/fonts/fonts/CARLITO-BOLD.TTF')
    fontDB.addApplicationFont(':/fonts/fonts/CARLITO-BOLDITALIC.TTF')
    fontDB.addApplicationFont(':/fonts/fonts/CARLITO-ITALIC.TTF')
    fontDB.addApplicationFont(':/fonts/fonts/CARLITO-REGULAR.TTF')
    fontDB.addApplicationFont(':/fonts/fonts/Carnivalee_Freakshow.ttf')
    fontDB.addApplicationFont(':/fonts/fonts/CENTAUR.TTF')
    fontDB.addApplicationFont(':/fonts/fonts/CENTURY.TTF')
    fontDB.addApplicationFont(':/fonts/fonts/COLONNA.TTF')
    fontDB.addApplicationFont(':/fonts/fonts/COMIC.TTF')
    fontDB.addApplicationFont(':/fonts/fonts/COMICBD.TTF')
    fontDB.addApplicationFont(':/fonts/fonts/COMICI.TTF')
    fontDB.addApplicationFont(':/fonts/fonts/COMICZ.TTF')
    fontDB.addApplicationFont(':/fonts/fonts/CONSTAN.TTF')
    fontDB.addApplicationFont(':/fonts/fonts/CONSTANB.TTF')
    fontDB.addApplicationFont(':/fonts/fonts/CONSTANI.TTF')
    fontDB.addApplicationFont(':/fonts/fonts/CONSTANZ.TTF')
    fontDB.addApplicationFont(':/fonts/fonts/COURE.FON')
    fontDB.addApplicationFont(':/fonts/fonts/C_BOX.TTF')
    fontDB.addApplicationFont(':/fonts/fonts/C_BOX_D.TTF')
    fontDB.addApplicationFont(':/fonts/fonts/Deloise.ttf')
    fontDB.addApplicationFont(':/fonts/fonts/Extra_Sales_Blank.ttf')
    fontDB.addApplicationFont(':/fonts/fonts/GABRIOLA.TTF')
    fontDB.addApplicationFont(':/fonts/fonts/GARA.TTF')
    fontDB.addApplicationFont(':/fonts/fonts/GARABD.TTF')
    fontDB.addApplicationFont(':/fonts/fonts/GARAIT.TTF')
    fontDB.addApplicationFont(':/fonts/fonts/GEORGIA.TTF')
    fontDB.addApplicationFont(':/fonts/fonts/GEORGIAB.TTF')
    fontDB.addApplicationFont(':/fonts/fonts/GEORGIAI.TTF')
    fontDB.addApplicationFont(':/fonts/fonts/GEORGIAZ.TTF')
    fontDB.addApplicationFont(':/fonts/fonts/GIDDYUPSTD.OTF')
    fontDB.addApplicationFont(':/fonts/fonts/Gold Plated.ttf')
    fontDB.addApplicationFont(':/fonts/fonts/GOTHIC.TTF')
    fontDB.addApplicationFont(':/fonts/fonts/GOTHICB.TTF')
    fontDB.addApplicationFont(':/fonts/fonts/GOTHICBI.TTF')
    fontDB.addApplicationFont(':/fonts/fonts/GOTHICI.TTF')
    fontDB.addApplicationFont(':/fonts/fonts/Graymalkin_Academy_Condensed.ttf')
    fontDB.addApplicationFont(':/fonts/fonts/Hadrianus_Demo.ttf')
    fontDB.addApplicationFont(':/fonts/fonts/Harry_Potter.ttf')
    fontDB.addApplicationFont(':/fonts/fonts/HTOWERT.TTF')
    fontDB.addApplicationFont(':/fonts/fonts/HTOWERTI.TTF')
    fontDB.addApplicationFont(':/fonts/fonts/InaiMathi.ttf')
    fontDB.addApplicationFont(':/fonts/fonts/Interceptor_Condensed.ttf')
    fontDB.addApplicationFont(':/fonts/fonts/JOKERMAN.TTF')
    fontDB.addApplicationFont(':/fonts/fonts/Jurassic-Regular.ttf')
    fontDB.addApplicationFont(':/fonts/fonts/LCALLIG.TTF')
    fontDB.addApplicationFont(':/fonts/fonts/Longdon_Decorative.ttf')
    fontDB.addApplicationFont(':/fonts/fonts/MERCEDES.TTF')
    fontDB.addApplicationFont(':/fonts/fonts/MTCORSVA.TTF')
    fontDB.addApplicationFont(':/fonts/fonts/N019043L.TTF')
    fontDB.addApplicationFont(':/fonts/fonts/N019044L.TTF')
    fontDB.addApplicationFont(':/fonts/fonts/N019063L.TTF')
    fontDB.addApplicationFont(':/fonts/fonts/N019064L.TTF')
    fontDB.addApplicationFont(':/fonts/fonts/Neuton_Cursive.ttf')
    fontDB.addApplicationFont(':/fonts/fonts/Puritan_Alternate_Bold.ttf')
    fontDB.addApplicationFont(':/fonts/fonts/ROSEWOODSTD-REGULAR.OTF')
    fontDB.addApplicationFont(':/fonts/fonts/ShouldveKnownShaded-Regular.ttf')
    fontDB.addApplicationFont(':/fonts/fonts/Simonetta-Black.ttf')
    fontDB.addApplicationFont(':/fonts/fonts/Simonetta-BlackItalic.ttf')
    fontDB.addApplicationFont(':/fonts/fonts/Simonetta-Italic.ttf')
    fontDB.addApplicationFont(':/fonts/fonts/Simonetta-Regular.ttf')
    fontDB.addApplicationFont(':/fonts/fonts/TAHOMA.TTF')
    fontDB.addApplicationFont(':/fonts/fonts/TAHOMABD.TTF')
    fontDB.addApplicationFont(':/fonts/fonts/TIMES.TTF')
    fontDB.addApplicationFont(':/fonts/fonts/TIMESBD.TTF')
    fontDB.addApplicationFont(':/fonts/fonts/TIMESBI.TTF')
    fontDB.addApplicationFont(':/fonts/fonts/TIMESI.TTF')
    fontDB.addApplicationFont(':/fonts/fonts/TRAJANPRO-BOLD.OTF')
    fontDB.addApplicationFont(':/fonts/fonts/TRAJANPRO-REGULAR.OTF')
    fontDB.addApplicationFont(':/fonts/fonts/Transformers_Movie.ttf')
    fontDB.addApplicationFont(':/fonts/fonts/trial.py')
    fontDB.addApplicationFont(':/fonts/fonts/UBUNTUMONO-B.TTF')
    fontDB.addApplicationFont(':/fonts/fonts/UBUNTUMONO-R.TTF')
    fontDB.addApplicationFont(':/fonts/fonts/UBUNTUMONO-RI.TTF')
    fontDB.addApplicationFont(':/fonts/fonts/ZaragozaPlain-Regular.ttf')
    fontDB.addApplicationFont(':/fonts/fonts/Zebrawood-Regular-Regular.ttf')
    fontDB.addApplicationFont(':/fonts/fonts/Zimbra-Bold.otf')
    fontDB.addApplicationFont(':/fonts/fonts/Zombie-A.ttf')
    fontDB.addApplicationFont(':/fonts/fonts/HARNGTON.TTF')


if __name__ == "__main__":
    import sys

    app = QtGui.QApplication(sys.argv)

    loginDialog = QtGui.QDialog()
    loginui = Ui_loginDialog()
    loginui.setupUi(loginDialog)

    MainWindow = QtGui.QMainWindow()
    ui = Ui_MainWindow()
    ui.setupUi(MainWindow)

    set_up_font()

    loginui.username = ''
    loginui.firstrun = True

    from time import sleep


    def checkuserpass():

        username = loginui.login_usernameLineEdit.displayText().strip()
        password = loginui.login_passwordLineEdit.text().strip()

        # username= 'ncc_editor'
        # password='nccindia'

        if username == 'ncc_editor' and password == 'nccindia':
            loginui.username = "ncc_editor"

        elif username == 'ncc_viewer' and password.strip() == '':
            loginui.username = "ncc_viewer"

        else:
            QtGui.QMessageBox.warning(loginDialog, "Login Failed", 'Username or Password Incorrect', 'OK')
            loginui.login_passwordLineEdit.clear()
            return

        if loginui.firstrun:
            loginDialog.hide()
            myobj = logic()
            MainWindow.show()
            loginui.loginpermission = myobj.login_permission
            loginui.login_passwordLineEdit.clear()
            loginui.login_usernameLineEdit.clear()
            loginDialog.close()
            app.processEvents()
            sleep(0.1)
            myobj.showtooltip("WELCOME")
        else:
            loginui.login_passwordLineEdit.clear()
            loginui.login_usernameLineEdit.clear()
            loginDialog.close()
            app.processEvents()
            sleep(0.1)
            loginui.loginpermission()


    if loginui.username == '':
        loginDialog.show()
        loginui.loginPushButton.clicked.connect(checkuserpass)

    object = QtGui.QGraphicsDropShadowEffect()
    object.setBlurRadius(5.2)

    button1 = QtGui.QPushButton()
    button1.setGraphicsEffect(object)

    button2 = QtGui.QPushButton()
    button2.setGraphicsEffect(object)

    # After some processing i want to disable the shadow for Button1

    object.setEnabled(False)

    # But this will disable shadow for both buttons !!!

    sys.exit(app.exec_())
